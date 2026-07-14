from __future__ import annotations

from calendar import monthrange
from datetime import datetime, timezone
from typing import List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.db.mongo import MongoClientManager
from app.models.user import UserPublic
from app.models.pm.reports import (
    MonthlyReportCreate, MonthlyReportPatch, MonthlyReportOut,
    ProjectBreakdown, PersonBreakdown, WorkItem, ReportStats, EditHistoryEntry,
)
from app.routers.auth import get_current_user
from app.routers.pm.report_agg import aggregate_period
from app.routers.pm.weekly_reports import (
    _require_pm, _history_entry, _user_name, _excel_response, _write_items_table,
    STATUS_KO, PRIORITY_KO, STATUS_ISSUE_KO,
)

router = APIRouter()


def _parse_items(raw: list) -> list[WorkItem]:
    return [WorkItem(**i) for i in raw]


def _parse_breakdown(raw: list, cls):
    result = []
    for d in raw:
        for key in ("completed", "in_progress", "delayed", "upcoming"):
            if key in d:
                d[key] = [WorkItem(**i) for i in d[key]]
        if "stats" in d and isinstance(d["stats"], dict):
            d["stats"] = ReportStats(**d["stats"])
        result.append(cls(**d))
    return result


def _doc_to_out(doc: dict) -> MonthlyReportOut:
    return MonthlyReportOut(
        id=str(doc["_id"]),
        report_year=doc["report_year"],
        report_month=doc["report_month"],
        title=doc["title"],
        department=doc.get("department"),
        by_project=_parse_breakdown(doc.get("by_project", []), ProjectBreakdown),
        by_person=_parse_breakdown(doc.get("by_person", []), PersonBreakdown),
        all_items=_parse_items(doc.get("all_items", [])),
        upcoming_items=_parse_items(doc.get("upcoming_items", [])),
        stats=ReportStats(**doc["stats"]) if doc.get("stats") else ReportStats(),
        admin_comment=doc.get("admin_comment"),
        created_by=str(doc["created_by"]),
        created_by_name=None,
        created_at=doc["created_at"],
        updated_by=str(doc["updated_by"]) if doc.get("updated_by") else None,
        updated_by_name=None,
        updated_at=doc["updated_at"],
        edit_history=[
            EditHistoryEntry(
                editor_id=str(e["editor_id"]),
                editor_name=e.get("editor_name"),
                action=e.get("action", "수정"),
                edited_at=e["edited_at"],
            )
            for e in doc.get("edit_history", [])
        ],
    )


async def _enrich(out: MonthlyReportOut, doc: dict) -> MonthlyReportOut:
    out.created_by_name = await _user_name(doc.get("created_by"))
    out.updated_by_name = await _user_name(doc.get("updated_by"))
    return out


def _month_range(year: int, month: int):
    last = monthrange(year, month)[1]
    start = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
    end   = datetime(year, month, last, 23, 59, 59, tzinfo=timezone.utc)
    return start, end


def _next_month_range(year: int, month: int):
    nm = month + 1
    ny = year
    if nm > 12:
        nm = 1; ny += 1
    return _month_range(ny, nm)


# ════════════════════════════════════════════════════════════════════
# CRUD
# ════════════════════════════════════════════════════════════════════

@router.get("", response_model=List[MonthlyReportOut])
async def list_monthly_reports(
    year:  Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    col = MongoClientManager.get_pm_monthly_reports_collection()
    q: dict = {"deleted_at": None}
    if year:  q["report_year"]  = year
    if month: q["report_month"] = month
    docs = await col.find(q).sort([("report_year", -1), ("report_month", -1)]).to_list(None)
    result = []
    for doc in docs:
        out = _doc_to_out(doc)
        out.created_by_name = await _user_name(doc.get("created_by"))
        result.append(out)
    return result


# NOTE: /export/list must be defined BEFORE /{report_id} to avoid FastAPI route conflict
@router.get("/export/list")
async def export_monthly_list(
    year:  Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    col = MongoClientManager.get_pm_monthly_reports_collection()
    q: dict = {"deleted_at": None}
    if year:  q["report_year"]  = year
    if month: q["report_month"] = month
    docs = await col.find(q).sort([("report_year", -1), ("report_month", -1)]).to_list(None)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "월간보고 목록"
    hf    = PatternFill("solid", fgColor="2F75B6")
    hfont = Font(color="FFFFFF", bold=True)

    headers = ["연도", "월", "부서", "제목", "총", "완료", "진행", "지연", "완료율", "작성자", "작성일"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center")

    for ri, doc in enumerate(docs, 2):
        s = doc.get("stats", {})
        ws.cell(ri, 1, doc["report_year"])
        ws.cell(ri, 2, f"{doc['report_month']}월")
        ws.cell(ri, 3, doc.get("department") or "")
        ws.cell(ri, 4, doc["title"])
        ws.cell(ri, 5, s.get("total", 0))
        ws.cell(ri, 6, s.get("completed", 0))
        ws.cell(ri, 7, s.get("in_progress", 0))
        ws.cell(ri, 8, s.get("delayed", 0))
        ws.cell(ri, 9, f"{s.get('completion_rate', 0)}%")
        ws.cell(ri, 10, await _user_name(doc.get("created_by")) or "")
        ws.cell(ri, 11, doc["created_at"].strftime("%Y-%m-%d"))

    for ci in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
    ws.column_dimensions["D"].width = 40

    return _excel_response(wb, f"월간보고목록_{year or '전체'}년.xlsx")


@router.get("/{report_id}", response_model=MonthlyReportOut)
async def get_monthly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    doc = await MongoClientManager.get_pm_monthly_reports_collection().find_one(
        {"_id": ObjectId(report_id), "deleted_at": None}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    return await _enrich(_doc_to_out(doc), doc)


@router.post("", response_model=MonthlyReportOut, status_code=201)
async def create_monthly_report(
    body: MonthlyReportCreate,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    col = MongoClientManager.get_pm_monthly_reports_collection()
    now = datetime.now(timezone.utc)

    start_dt, end_dt = _month_range(body.report_year, body.report_month)
    ns, ne = _next_month_range(body.report_year, body.report_month)
    all_items, upcoming_items, by_project, by_person, stats = await aggregate_period(
        start_dt, end_dt, ns, ne
    )

    result = await col.insert_one({
        "report_year":    body.report_year,
        "report_month":   body.report_month,
        "title":          body.title,
        "department":     body.department,
        "by_project":     [p.model_dump() for p in by_project],
        "by_person":      [p.model_dump() for p in by_person],
        "all_items":      [i.model_dump() for i in all_items],
        "upcoming_items": [i.model_dump() for i in upcoming_items],
        "stats":          stats.model_dump(),
        "admin_comment":  body.admin_comment,
        "created_by":     ObjectId(current_user.id),
        "updated_by":     None,
        "created_at":     now,
        "updated_at":     now,
        "deleted_at":     None,
        "edit_history":   [_history_entry(current_user, "생성")],
    })
    doc = await col.find_one({"_id": result.inserted_id})
    return await _enrich(_doc_to_out(doc), doc)


@router.patch("/{report_id}", response_model=MonthlyReportOut)
async def patch_monthly_report(
    report_id: str,
    body: MonthlyReportPatch,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    col = MongoClientManager.get_pm_monthly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    patch = body.model_dump(exclude_unset=True)
    patch["updated_by"] = ObjectId(current_user.id)
    patch["updated_at"] = datetime.now(timezone.utc)

    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)},
        {"$set": patch, "$push": {"edit_history": _history_entry(current_user, "수정")}},
        return_document=True,
    )
    return await _enrich(_doc_to_out(updated), updated)


@router.post("/{report_id}/refresh", response_model=MonthlyReportOut)
async def refresh_monthly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    """이슈 데이터를 다시 집계하여 보고서 내용을 최신화합니다."""
    _require_pm(current_user)
    col = MongoClientManager.get_pm_monthly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    start_dt, end_dt = _month_range(doc["report_year"], doc["report_month"])
    ns, ne = _next_month_range(doc["report_year"], doc["report_month"])
    all_items, upcoming_items, by_project, by_person, stats = await aggregate_period(
        start_dt, end_dt, ns, ne
    )
    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)},
        {
            "$set": {
                "by_project":     [p.model_dump() for p in by_project],
                "by_person":      [p.model_dump() for p in by_person],
                "all_items":      [i.model_dump() for i in all_items],
                "upcoming_items": [i.model_dump() for i in upcoming_items],
                "stats":          stats.model_dump(),
                "updated_by":     ObjectId(current_user.id),
                "updated_at":     datetime.now(timezone.utc),
            },
            "$push": {"edit_history": _history_entry(current_user, "데이터 새로고침")},
        },
        return_document=True,
    )
    return await _enrich(_doc_to_out(updated), updated)


@router.delete("/{report_id}", status_code=204)
async def delete_monthly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    result = await MongoClientManager.get_pm_monthly_reports_collection().update_one(
        {"_id": ObjectId(report_id), "deleted_at": None},
        {"$set": {"deleted_at": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")


@router.get("/{report_id}/export")
async def export_monthly_detail(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_pm(current_user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    doc = await MongoClientManager.get_pm_monthly_reports_collection().find_one(
        {"_id": ObjectId(report_id), "deleted_at": None}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    r = _doc_to_out(doc)
    r.created_by_name = await _user_name(doc.get("created_by"))

    hf    = PatternFill("solid", fgColor="2F75B6")
    hfont = Font(color="FFFFFF", bold=True)
    alt   = PatternFill("solid", fgColor="DCE6F1")
    bold  = Font(bold=True)

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "개요"
    ws1["A1"] = r.title; ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")
    for i, (lbl, val) in enumerate([
        ("보고 기간", f"{r.report_year}년 {r.report_month}월"),
        ("부서", r.department or ""), ("작성자", r.created_by_name or ""),
    ], 3):
        ws1.cell(i, 1, lbl).font = bold; ws1.cell(i, 2, val)

    row = 8
    ws1.cell(row, 1, "업무 통계").font = Font(bold=True, size=12); row += 1
    for lbl, val in [("총 업무", r.stats.total), ("완료", r.stats.completed),
                     ("진행 중", r.stats.in_progress), ("지연", r.stats.delayed),
                     ("완료율", f"{r.stats.completion_rate}%")]:
        ws1.cell(row, 1, lbl).font = bold; ws1.cell(row, 2, val); row += 1

    if r.admin_comment:
        row += 1
        ws1.cell(row, 1, "관리자 코멘트").font = Font(bold=True, size=12); row += 1
        ws1.cell(row, 1, r.admin_comment)

    for ws_title, breakdowns, next_label in [
        ("프로젝트별", r.by_project, "차월 계획"),
        ("개인별", r.by_person, "차월 계획"),
    ]:
        wsn = wb.create_sheet(ws_title); rown = 1
        for pb in breakdowns:
            name = pb.project_name if hasattr(pb, "project_name") else pb.user_name
            wsn.cell(rown, 1, f"▶ {name}").font = Font(bold=True, size=12)
            wsn.merge_cells(f"A{rown}:K{rown}"); rown += 1
            wsn.cell(rown, 1, f"완료 {pb.stats.completed} / 진행 {pb.stats.in_progress} / 지연 {pb.stats.delayed} / 완료율 {pb.stats.completion_rate}%")
            rown += 2
            for sec, items in [("완료", pb.completed), ("진행 중", pb.in_progress), ("지연", pb.delayed), (next_label, pb.upcoming)]:
                if not items: continue
                wsn.cell(rown, 1, sec).font = bold; rown += 1
                rown = _write_items_table(wsn, items, rown, hf, hfont, alt); rown += 1
            rown += 1

    ws4 = wb.create_sheet("전체 업무")
    _write_items_table(ws4, r.all_items, 1, hf, hfont, alt)

    if r.upcoming_items:
        ws5 = wb.create_sheet("차월 계획")
        _write_items_table(ws5, r.upcoming_items, 1, hf, hfont, alt)

    for ws in wb.worksheets:
        for ci in range(1, 12):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18
        ws.column_dimensions["B"].width = 40

    return _excel_response(wb, f"월간보고_{r.report_year}년_{r.report_month:02d}월.xlsx")
