from __future__ import annotations

from datetime import datetime, timedelta, timezone
from typing import List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.db.mongo import MongoClientManager
from app.models.user import UserPublic
from app.models.pm.reports import (
    WeeklyReportCreate, WeeklyReportPatch, WeeklyReportOut,
    WeeklyReportStatusPatch, ManualItemCreate, ManualItemPatch, ManualItemOut,
    ProjectBreakdown, PersonBreakdown, WorkItem, ReportStats,
)
from app.routers.auth import get_current_user
from app.routers.pm.report_agg import aggregate_period

router = APIRouter()

STATUS_KO       = {"DRAFT": "초안", "REVIEWING": "검토중", "CONFIRMED": "확정", "PUBLISHED": "게시됨", "ARCHIVED": "보관됨"}
PRIORITY_KO     = {"LOWEST": "최하", "LOW": "낮음", "MEDIUM": "중간", "HIGH": "높음", "HIGHEST": "최고"}
STATUS_ISSUE_KO = {
    "BACKLOG": "백로그", "TODO": "할 일",
    "IN_PROGRESS": "진행 중", "IN_REVIEW": "검토 중", "DONE": "완료",
}


def _require_admin(user: UserPublic):
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")


async def _user_name(uid) -> Optional[str]:
    if not uid:
        return None
    u = await MongoClientManager.get_users_collection().find_one(
        {"_id": ObjectId(str(uid))}, {"full_name": 1, "email": 1}
    )
    return (u.get("full_name") or u.get("email")) if u else None


def _parse_items(raw: list) -> list[WorkItem]:
    return [WorkItem(**i) for i in raw]


def _parse_manual_items(raw: list) -> list[ManualItemOut]:
    result = []
    for item in raw:
        result.append(ManualItemOut(
            id=str(item["_id"]),
            section=item.get("section", ""),
            title=item.get("title", ""),
            owner=item.get("owner"),
            linked_sr_id=item.get("linked_sr_id"),
            linked_issue_id=item.get("linked_issue_id"),
            include_in_report=item.get("include_in_report", True),
            sort_order=item.get("sort_order", 0),
            category=item.get("category"),
            content=item.get("content"),
            agenda_status=item.get("agenda_status"),
            item_type=item.get("item_type"),
            impact=item.get("impact"),
            action_plan=item.get("action_plan"),
            background=item.get("background"),
            options=item.get("options"),
            requested_decision=item.get("requested_decision"),
            desired_date=item.get("desired_date"),
            created_by=str(item["created_by"]),
            created_at=item["created_at"],
            updated_by=str(item["updated_by"]) if item.get("updated_by") else None,
            updated_at=item["updated_at"],
        ))
    return result


def _parse_breakdown(raw: list, cls):
    result = []
    for d in raw:
        for key in ("completed", "in_progress", "delayed", "upcoming"):
            if key in d:
                d[key] = [WorkItem(**i) for i in d[key]]
        if "stats" in d and isinstance(d["stats"], dict):
            d["stats"] = ReportStats(**d["stats"])
        result.append(cls(**d))
    return result


def _doc_to_out(doc: dict) -> WeeklyReportOut:
    return WeeklyReportOut(
        id=str(doc["_id"]),
        report_year=doc["report_year"],
        report_week=doc["report_week"],
        start_date=doc["start_date"],
        end_date=doc["end_date"],
        title=doc["title"],
        department=doc.get("department"),
        status=doc.get("status", "DRAFT"),
        by_project=_parse_breakdown(doc.get("by_project", []), ProjectBreakdown),
        by_person=_parse_breakdown(doc.get("by_person", []), PersonBreakdown),
        all_items=_parse_items(doc.get("all_items", [])),
        upcoming_items=_parse_items(doc.get("upcoming_items", [])),
        stats=ReportStats(**doc["stats"]) if doc.get("stats") else ReportStats(),
        manual_items=_parse_manual_items(doc.get("manual_items", [])),
        admin_comment=doc.get("admin_comment"),
        created_by=str(doc["created_by"]),
        created_by_name=None,
        created_at=doc["created_at"],
        updated_by=str(doc["updated_by"]) if doc.get("updated_by") else None,
        updated_by_name=None,
        updated_at=doc["updated_at"],
        confirmed_by=str(doc["confirmed_by"]) if doc.get("confirmed_by") else None,
        confirmed_at=doc.get("confirmed_at"),
    )


async def _enrich(out: WeeklyReportOut, doc: dict) -> WeeklyReportOut:
    out.created_by_name = await _user_name(doc.get("created_by"))
    out.updated_by_name = await _user_name(doc.get("updated_by"))
    return out


def _next_week(start_dt: datetime, end_dt: datetime):
    delta = end_dt - start_dt + timedelta(seconds=1)
    ns = end_dt + timedelta(seconds=1)
    return ns, ns + delta - timedelta(seconds=1)


def _excel_response(wb, filename: str) -> StreamingResponse:
    import io
    from urllib.parse import quote
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _write_items_table(ws, items: list, start_row: int, hf, hfont, alt):
    import openpyxl
    from openpyxl.styles import Alignment
    headers = ["번호", "업무명", "에픽", "스프린트", "담당자", "우선순위", "상태", "시작일", "마감일", "SP", "지연"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    for ri, item in enumerate(items, start_row + 1):
        vals = [
            f"{item['project_name']}-{item['issue_number']}" if isinstance(item, dict) else f"{item.project_name}-{item.issue_number}",
            item['title'] if isinstance(item, dict) else item.title,
            (item.get('epic_title') or "") if isinstance(item, dict) else (item.epic_title or ""),
            (item.get('sprint_name') or "") if isinstance(item, dict) else (item.sprint_name or ""),
            (item.get('assignee_name') or "미지정") if isinstance(item, dict) else (item.assignee_name or "미지정"),
            PRIORITY_KO.get(item['priority'] if isinstance(item, dict) else item.priority, ""),
            STATUS_ISSUE_KO.get(item['status'] if isinstance(item, dict) else item.status, ""),
            "",  # start_date handled below
            "",  # due_date handled below
            (item.get('story_points') or "") if isinstance(item, dict) else (item.story_points or ""),
            "지연" if (item.get('is_delayed') if isinstance(item, dict) else item.is_delayed) else "",
        ]
        sd = item.get('start_date') if isinstance(item, dict) else item.start_date
        dd = item.get('due_date')   if isinstance(item, dict) else item.due_date
        vals[7] = sd.strftime("%Y-%m-%d") if sd else ""
        vals[8] = dd.strftime("%Y-%m-%d") if dd else ""
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if ri % 2 == 0:
                c.fill = alt
    return start_row + 1 + len(items)


# ════════════════════════════════════════════════════════════════════
# CRUD
# ════════════════════════════════════════════════════════════════════

@router.get("", response_model=List[WeeklyReportOut])
async def list_weekly_reports(
    year: Optional[int] = Query(None),
    week: Optional[int] = Query(None),
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    q: dict = {"deleted_at": None}
    if year: q["report_year"] = year
    if week: q["report_week"] = week
    docs = await col.find(q).sort([("report_year", -1), ("report_week", -1)]).to_list(None)
    result = []
    for doc in docs:
        out = _doc_to_out(doc)
        out.created_by_name = await _user_name(doc.get("created_by"))
        result.append(out)
    return result


# NOTE: /export/list must be defined BEFORE /{report_id} to avoid FastAPI route conflict
@router.get("/export/list")
async def export_weekly_list(
    year: Optional[int] = Query(None),
    week: Optional[int] = Query(None),
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    col = MongoClientManager.get_pm_weekly_reports_collection()
    q: dict = {"deleted_at": None}
    if year: q["report_year"] = year
    if week: q["report_week"] = week
    docs = await col.find(q).sort([("report_year", -1), ("report_week", -1)]).to_list(None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주간보고 목록"
    hf    = PatternFill("solid", fgColor="2F75B6")
    hfont = Font(color="FFFFFF", bold=True)

    headers = ["연도", "주차", "보고 기간", "부서", "제목", "총", "완료", "진행", "지연", "완료율"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center")

    for ri, doc in enumerate(docs, 2):
        s = doc.get("stats", {})
        ws.cell(ri, 1, doc["report_year"])
        ws.cell(ri, 2, f"{doc['report_week']}주")
        ws.cell(ri, 3, f"{doc['start_date'].strftime('%m/%d')}~{doc['end_date'].strftime('%m/%d')}")
        ws.cell(ri, 4, doc.get("department") or "")
        ws.cell(ri, 5, doc["title"])
        ws.cell(ri, 6, s.get("total", 0))
        ws.cell(ri, 7, s.get("completed", 0))
        ws.cell(ri, 8, s.get("in_progress", 0))
        ws.cell(ri, 9, s.get("delayed", 0))
        ws.cell(ri, 10, f"{s.get('completion_rate', 0)}%")

    for ci in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
    ws.column_dimensions["E"].width = 40

    return _excel_response(wb, f"주간보고목록_{year or '전체'}년.xlsx")


@router.get("/{report_id}", response_model=WeeklyReportOut)
async def get_weekly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    doc = await MongoClientManager.get_pm_weekly_reports_collection().find_one(
        {"_id": ObjectId(report_id), "deleted_at": None}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    return await _enrich(_doc_to_out(doc), doc)


@router.post("", response_model=WeeklyReportOut, status_code=201)
async def create_weekly_report(
    body: WeeklyReportCreate,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    now = datetime.now(timezone.utc)

    ns, ne = _next_week(body.start_date, body.end_date)
    all_items, upcoming_items, by_project, by_person, stats = await aggregate_period(
        body.start_date, body.end_date, ns, ne
    )

    result = await col.insert_one({
        "report_year":    body.report_year,
        "report_week":    body.report_week,
        "start_date":     body.start_date,
        "end_date":       body.end_date,
        "title":          body.title,
        "department":     body.department,
        "status":         "DRAFT",
        "by_project":     [p.model_dump() for p in by_project],
        "by_person":      [p.model_dump() for p in by_person],
        "all_items":      [i.model_dump() for i in all_items],
        "upcoming_items": [i.model_dump() for i in upcoming_items],
        "stats":          stats.model_dump(),
        "manual_items":   [],
        "admin_comment":  body.admin_comment,
        "created_by":     ObjectId(current_user.id),
        "updated_by":     None,
        "created_at":     now,
        "updated_at":     now,
        "deleted_at":     None,
        "confirmed_by":   None,
        "confirmed_at":   None,
    })
    doc = await col.find_one({"_id": result.inserted_id})
    return await _enrich(_doc_to_out(doc), doc)


@router.patch("/{report_id}", response_model=WeeklyReportOut)
async def patch_weekly_report(
    report_id: str,
    body: WeeklyReportPatch,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    patch = body.model_dump(exclude_unset=True)
    patch["updated_by"] = ObjectId(current_user.id)
    patch["updated_at"] = datetime.now(timezone.utc)

    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)}, {"$set": patch}, return_document=True
    )
    return await _enrich(_doc_to_out(updated), updated)


@router.post("/{report_id}/refresh", response_model=WeeklyReportOut)
async def refresh_weekly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    """이슈 데이터를 다시 집계하여 보고서 내용을 최신화합니다."""
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    ns, ne = _next_week(doc["start_date"], doc["end_date"])
    all_items, upcoming_items, by_project, by_person, stats = await aggregate_period(
        doc["start_date"], doc["end_date"], ns, ne
    )
    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)},
        {"$set": {
            "by_project":     [p.model_dump() for p in by_project],
            "by_person":      [p.model_dump() for p in by_person],
            "all_items":      [i.model_dump() for i in all_items],
            "upcoming_items": [i.model_dump() for i in upcoming_items],
            "stats":          stats.model_dump(),
            "updated_by":     ObjectId(current_user.id),
            "updated_at":     datetime.now(timezone.utc),
        }},
        return_document=True,
    )
    return await _enrich(_doc_to_out(updated), updated)


# ════════════════════════════════════════════════════════════════════
# 상태 관리
# ════════════════════════════════════════════════════════════════════

@router.patch("/{report_id}/status", response_model=WeeklyReportOut)
async def change_report_status(
    report_id: str,
    body: WeeklyReportStatusPatch,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    allowed = {"DRAFT", "REVIEWING", "CONFIRMED"}
    if body.status not in allowed:
        raise HTTPException(status_code=400, detail=f"유효하지 않은 상태값입니다. ({', '.join(allowed)})")
    col = MongoClientManager.get_pm_weekly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    now = datetime.now(timezone.utc)
    update: dict = {"status": body.status, "updated_by": ObjectId(current_user.id), "updated_at": now}
    if body.status == "CONFIRMED":
        update["confirmed_by"] = ObjectId(current_user.id)
        update["confirmed_at"] = now
    elif body.status != "CONFIRMED":
        # 확정 해제 시 초기화
        update["confirmed_by"] = None
        update["confirmed_at"] = None

    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)}, {"$set": update}, return_document=True
    )
    return await _enrich(_doc_to_out(updated), updated)


# ════════════════════════════════════════════════════════════════════
# 수기 항목 CRUD
# ════════════════════════════════════════════════════════════════════

@router.post("/{report_id}/items", response_model=WeeklyReportOut, status_code=201)
async def add_manual_item(
    report_id: str,
    body: ManualItemCreate,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    if doc.get("status") == "CONFIRMED":
        raise HTTPException(status_code=403, detail="확정된 보고서는 수정할 수 없습니다.")

    now = datetime.now(timezone.utc)
    item_data = body.model_dump(exclude_none=True)
    item_data["_id"] = ObjectId()
    item_data["created_by"] = ObjectId(current_user.id)
    item_data["created_at"] = now
    item_data["updated_by"] = None
    item_data["updated_at"] = now

    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)},
        {
            "$push": {"manual_items": item_data},
            "$set": {"updated_at": now, "updated_by": ObjectId(current_user.id)},
        },
        return_document=True,
    )
    return await _enrich(_doc_to_out(updated), updated)


@router.patch("/{report_id}/items/{item_id}", response_model=WeeklyReportOut)
async def update_manual_item(
    report_id: str,
    item_id: str,
    body: ManualItemPatch,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    if doc.get("status") == "CONFIRMED":
        raise HTTPException(status_code=403, detail="확정된 보고서는 수정할 수 없습니다.")

    now = datetime.now(timezone.utc)
    patch = body.model_dump(exclude_unset=True)
    patch["updated_by"] = ObjectId(current_user.id)
    patch["updated_at"] = now

    set_fields: dict = {f"manual_items.$[item].{k}": v for k, v in patch.items()}
    set_fields["updated_at"] = now
    set_fields["updated_by"] = ObjectId(current_user.id)

    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)},
        {"$set": set_fields},
        array_filters=[{"item._id": ObjectId(item_id)}],
        return_document=True,
    )
    if not updated:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다.")
    return await _enrich(_doc_to_out(updated), updated)


@router.delete("/{report_id}/items/{item_id}", response_model=WeeklyReportOut)
async def delete_manual_item(
    report_id: str,
    item_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    col = MongoClientManager.get_pm_weekly_reports_collection()
    doc = await col.find_one({"_id": ObjectId(report_id), "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    if doc.get("status") == "CONFIRMED":
        raise HTTPException(status_code=403, detail="확정된 보고서는 수정할 수 없습니다.")

    now = datetime.now(timezone.utc)
    updated = await col.find_one_and_update(
        {"_id": ObjectId(report_id)},
        {
            "$pull": {"manual_items": {"_id": ObjectId(item_id)}},
            "$set": {"updated_at": now, "updated_by": ObjectId(current_user.id)},
        },
        return_document=True,
    )
    if not updated:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    return await _enrich(_doc_to_out(updated), updated)


# ════════════════════════════════════════════════════════════════════
# 보고서 미리보기 (텍스트)
# ════════════════════════════════════════════════════════════════════

@router.get("/{report_id}/preview")
async def preview_weekly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    doc = await MongoClientManager.get_pm_weekly_reports_collection().find_one(
        {"_id": ObjectId(report_id), "deleted_at": None}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    r = _doc_to_out(doc)
    r.created_by_name = await _user_name(doc.get("created_by"))

    lines: list[str] = []
    lines.append(f"# {r.title}")
    lines.append(f"보고 기간: {r.start_date.strftime('%Y-%m-%d')} ~ {r.end_date.strftime('%Y-%m-%d')} ({r.report_year}년 {r.report_week}주차)")
    if r.department:
        lines.append(f"부서: {r.department}")
    lines.append(f"작성자: {r.created_by_name or '-'}")
    lines.append(f"상태: {STATUS_KO.get(r.status, r.status)}")
    lines.append("")

    # 1. 금주 완료 업무
    completed = [i for i in r.all_items if i.status == "DONE"]
    if completed:
        lines.append("## 1. 금주 완료 업무")
        for item in completed:
            delayed_mark = " ⚠️지연" if item.is_delayed else ""
            lines.append(f"- [{item.project_name}-{item.issue_number}] {item.title}{delayed_mark}")
        lines.append("")

    # 2. 진행 중 업무
    in_prog = [i for i in r.all_items if i.status in ("IN_PROGRESS", "IN_REVIEW")]
    if in_prog:
        lines.append("## 2. 진행 중 업무")
        for item in in_prog:
            delayed_mark = " ⚠️지연" if item.is_delayed else ""
            due_str = f" — {item.due_date.strftime('%m/%d') if item.due_date else ''}"
            lines.append(f"- [{item.project_name}-{item.issue_number}] {item.title}{due_str}{delayed_mark}")
        lines.append("")

    # 3. 차주 계획
    if r.upcoming_items:
        lines.append("## 3. 차주 계획")
        for item in r.upcoming_items:
            lines.append(f"- [{item.project_name}-{item.issue_number}] {item.title}")
        lines.append("")

    # 4. 주요 안건
    agendas = [i for i in r.manual_items if i.section == "MAIN_AGENDA" and i.include_in_report]
    if agendas:
        lines.append("## 4. 주요 안건")
        for item in sorted(agendas, key=lambda x: x.sort_order):
            cat = f"[{item.category}] " if item.category else ""
            lines.append(f"- {cat}{item.title}")
            if item.content:
                lines.append(f"  - 내용: {item.content}")
            if item.agenda_status:
                lines.append(f"  - 진행 상태: {item.agenda_status}")
            if item.owner:
                lines.append(f"  - 담당자: {item.owner}")
        lines.append("")

    # 5. 특이사항 및 리스크
    risks = [i for i in r.manual_items if i.section == "ISSUE_RISK" and i.include_in_report]
    if risks:
        lines.append("## 5. 특이사항 및 리스크")
        for item in sorted(risks, key=lambda x: x.sort_order):
            type_label = f"[{item.item_type}/{item.impact}] " if item.item_type else ""
            lines.append(f"- {type_label}{item.title}")
            if item.content:
                lines.append(f"  - 내용: {item.content}")
            if item.action_plan:
                lines.append(f"  - 대응 방안: {item.action_plan}")
            if item.owner:
                lines.append(f"  - 담당자: {item.owner}")
        lines.append("")

    # 6. 결정 필요 사항
    decisions = [i for i in r.manual_items if i.section == "DECISION_REQUIRED" and i.include_in_report]
    if decisions:
        lines.append("## 6. 결정 필요 사항")
        for item in sorted(decisions, key=lambda x: x.sort_order):
            lines.append(f"- {item.title}")
            if item.background:
                lines.append(f"  - 배경: {item.background}")
            if item.options:
                lines.append(f"  - 선택지: {item.options}")
            if item.requested_decision:
                lines.append(f"  - 요청 결정: {item.requested_decision}")
            if item.desired_date:
                lines.append(f"  - 희망 결정일: {item.desired_date.strftime('%Y-%m-%d')}")
            if item.owner:
                lines.append(f"  - 담당자: {item.owner}")
        lines.append("")

    if r.admin_comment:
        lines.append("## 관리자 코멘트")
        lines.append(r.admin_comment)
        lines.append("")

    return {"text": "\n".join(lines)}


@router.delete("/{report_id}", status_code=204)
async def delete_weekly_report(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    result = await MongoClientManager.get_pm_weekly_reports_collection().update_one(
        {"_id": ObjectId(report_id), "deleted_at": None},
        {"$set": {"deleted_at": datetime.now(timezone.utc)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")


@router.get("/{report_id}/export")
async def export_weekly_detail(
    report_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    _require_admin(current_user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    doc = await MongoClientManager.get_pm_weekly_reports_collection().find_one(
        {"_id": ObjectId(report_id), "deleted_at": None}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")

    r = _doc_to_out(doc)
    r.created_by_name = await _user_name(doc.get("created_by"))

    hf    = PatternFill("solid", fgColor="2F75B6")
    hfont = Font(color="FFFFFF", bold=True)
    alt   = PatternFill("solid", fgColor="DCE6F1")
    bold  = Font(bold=True)

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "개요"
    ws1["A1"] = r.title; ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")
    for i, (lbl, val) in enumerate([
        ("보고 기간", f"{r.report_year}년 {r.report_week}주차 ({r.start_date.strftime('%m/%d')}~{r.end_date.strftime('%m/%d')})"),
        ("부서", r.department or ""), ("작성자", r.created_by_name or ""),
    ], 3):
        ws1.cell(i, 1, lbl).font = bold; ws1.cell(i, 2, val)

    row = 8
    ws1.cell(row, 1, "업무 통계").font = Font(bold=True, size=12); row += 1
    for lbl, val in [("총 업무", r.stats.total), ("완료", r.stats.completed),
                     ("진행 중", r.stats.in_progress), ("지연", r.stats.delayed),
                     ("완료율", f"{r.stats.completion_rate}%")]:
        ws1.cell(row, 1, lbl).font = bold; ws1.cell(row, 2, val); row += 1

    if r.admin_comment:
        row += 1
        ws1.cell(row, 1, "관리자 코멘트").font = Font(bold=True, size=12); row += 1
        ws1.cell(row, 1, r.admin_comment)

    # 프로젝트별
    ws2 = wb.create_sheet("프로젝트별"); row2 = 1
    for pb in r.by_project:
        ws2.cell(row2, 1, f"▶ {pb.project_name}").font = Font(bold=True, size=12)
        ws2.merge_cells(f"A{row2}:K{row2}"); row2 += 1
        ws2.cell(row2, 1, f"완료 {pb.stats.completed} / 진행 {pb.stats.in_progress} / 지연 {pb.stats.delayed} / 완료율 {pb.stats.completion_rate}%")
        row2 += 2
        for sec, items in [("완료", pb.completed), ("진행 중", pb.in_progress), ("지연", pb.delayed), ("차주 계획", pb.upcoming)]:
            if not items: continue
            ws2.cell(row2, 1, sec).font = bold; row2 += 1
            row2 = _write_items_table(ws2, items, row2, hf, hfont, alt); row2 += 1
        row2 += 1

    # 개인별
    ws3 = wb.create_sheet("개인별"); row3 = 1
    for pb in r.by_person:
        ws3.cell(row3, 1, f"▶ {pb.user_name}").font = Font(bold=True, size=12)
        ws3.merge_cells(f"A{row3}:K{row3}"); row3 += 1
        ws3.cell(row3, 1, f"완료 {pb.stats.completed} / 진행 {pb.stats.in_progress} / 지연 {pb.stats.delayed} / 완료율 {pb.stats.completion_rate}%")
        row3 += 2
        for sec, items in [("완료", pb.completed), ("진행 중", pb.in_progress), ("지연", pb.delayed), ("차주 계획", pb.upcoming)]:
            if not items: continue
            ws3.cell(row3, 1, sec).font = bold; row3 += 1
            row3 = _write_items_table(ws3, items, row3, hf, hfont, alt); row3 += 1
        row3 += 1

    ws4 = wb.create_sheet("전체 업무")
    _write_items_table(ws4, r.all_items, 1, hf, hfont, alt)

    if r.upcoming_items:
        ws5 = wb.create_sheet("차주 계획")
        _write_items_table(ws5, r.upcoming_items, 1, hf, hfont, alt)

    for ws in wb.worksheets:
        for ci in range(1, 12):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18
        ws.column_dimensions["B"].width = 40

    return _excel_response(wb, f"{r.title}.xlsx")
