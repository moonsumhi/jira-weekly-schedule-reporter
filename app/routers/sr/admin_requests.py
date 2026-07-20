"""관리자/처리자용 SR API: 전체 목록, 검토, 배정, 상태 변경, Excel 다운로드, 통계."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.db.mongo import MongoClientManager
from app.models.user import UserPublic
from app.models.sr.service_request import (
    SROut, SRListItem, SRListPage, SRPatch, SRInlinePatch,
    SRReview, SRAssign, SRStatusChange, SRDueDateChange,
    SRStats, SR_STATUS_LABEL, REQUEST_TYPE_LABEL, SR_PRIORITY_LABEL,
)
from app.routers.auth import get_current_user
from app.services.sr.sr_service import (
    get_sr_or_404, record_sr_history, record_status_history,
    record_due_date_history, sr_to_out, require_sr_operator,
    require_sr_manager, require_sr_admin, compute_is_delayed,
    is_sr_operator,
)
from app.services.notification_service import create_notification, notify_users

router = APIRouter()


def _user_label(user: UserPublic) -> str:
    return user.full_name or user.email


# ── 전체 SR 목록 ──────────────────────────────────────────────────────

@router.get("", response_model=SRListPage)
async def list_all_srs(
    status: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    requester_department: Optional[str] = Query(None),
    requester_name: Optional[str] = Query(None),
    related_system: Optional[str] = Query(None),
    assignee_id: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    is_urgent: Optional[bool] = Query(None),
    is_delayed: Optional[bool] = Query(None),
    my_assigned: Optional[bool] = Query(None),
    desired_due_from: Optional[str] = Query(None),
    desired_due_to: Optional[str] = Query(None),
    planned_due_from: Optional[str] = Query(None),
    planned_due_to: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    sort_by: Optional[str] = Query(None),
    descending: bool = Query(True),
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_operator(current_user)
    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]

    _SORT_FIELDS = {
        "created_at", "desired_due_date", "sr_no",
        "requester_name", "requester_department", "priority", "status",
    }
    sort_field = sort_by if sort_by in _SORT_FIELDS else "created_at"
    sort_dir   = -1 if descending else 1

    q: dict = {"deleted_at": None}
    # 자신에게 배정된 SR만 보기 (operator 전용 필터)
    if my_assigned:
        q["assignee_id"] = ObjectId(current_user.id)
    if status:
        q["status"] = status
    else:
        # 특정 status 필터 없이 전체 조회 시 임시저장(DRAFT) 제외
        q["status"] = {"$ne": "DRAFT"}
    if request_type:
        q["request_type"] = request_type
    if requester_department:
        q["requester_department"] = {"$regex": requester_department, "$options": "i"}
    if requester_name:
        q["requester_name"] = {"$regex": requester_name, "$options": "i"}
    if related_system:
        q["related_system"] = {"$regex": related_system, "$options": "i"}
    if assignee_id:
        q["assignee_id"] = ObjectId(assignee_id)
    if priority:
        q["priority"] = priority
    if is_urgent is not None:
        q["is_urgent"] = is_urgent

    def _dt(s: str) -> datetime:
        return datetime.fromisoformat(s).replace(tzinfo=timezone.utc)

    if desired_due_from or desired_due_to:
        df: dict = {}
        if desired_due_from:
            df["$gte"] = _dt(desired_due_from)
        if desired_due_to:
            df["$lte"] = _dt(desired_due_to)
        q["desired_due_date"] = df

    if planned_due_from or planned_due_to:
        pf: dict = {}
        if planned_due_from:
            pf["$gte"] = _dt(planned_due_from)
        if planned_due_to:
            pf["$lte"] = _dt(planned_due_to)
        q["planned_due_date"] = pf

    # is_delayed 필터는 Python 연산이므로 페이지네이션 전에 전체 조회 후 필터링
    if is_delayed is not None:
        all_docs = await col.find(q).sort(sort_field, sort_dir).to_list(None)
        outs = [sr_to_out(d) for d in all_docs]
        outs = [o for o in outs if o["is_delayed"] == is_delayed]
        page = outs[skip: skip + limit]
        return SRListPage(items=[SRListItem(**o) for o in page], total=len(outs))

    total = await col.count_documents(q)
    docs = await col.find(q).sort(sort_field, sort_dir).skip(skip).limit(limit).to_list(None)
    outs = [sr_to_out(d) for d in docs]
    return SRListPage(items=[SRListItem(**o) for o in outs], total=total)


# ── SR 상세 (관리자용) ────────────────────────────────────────────────

@router.get("/{sr_id}", response_model=SROut)
async def get_sr_admin(
    sr_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_operator(current_user)
    doc = await get_sr_or_404(sr_id)
    return SROut(**sr_to_out(doc))


# ── 인라인 필드 수정 (manager 이상) ─────────────────────────────────────

@router.patch("/{sr_id}", response_model=SROut)
async def patch_sr_inline(
    sr_id: str,
    body: SRInlinePatch,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_manager(current_user)
    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    doc = await get_sr_or_404(sr_id)
    now = datetime.now(timezone.utc)
    updates: dict = {"updated_at": now, "updated_by": _user_label(current_user)}

    patch = body.model_dump(exclude_none=True)
    for field, value in patch.items():
        old_val = doc.get(field)
        updates[field] = ObjectId(value) if field == "assignee_id" else value
        if str(old_val) != str(value):
            await record_sr_history(sr_id, f"FIELD_CHANGE:{field}", str(old_val), str(value), _user_label(current_user))

    await col.update_one({"_id": ObjectId(sr_id)}, {"$set": updates})
    updated = await col.find_one({"_id": ObjectId(sr_id)})
    return SROut(**sr_to_out(updated))


# ── SR 수정 (관리자) ──────────────────────────────────────────────────

@router.put("/{sr_id}", response_model=SROut)
async def update_sr_admin(
    sr_id: str,
    body: SRPatch,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_admin(current_user)
    doc = await get_sr_or_404(sr_id)
    now = datetime.now(timezone.utc)
    updates: dict = {"updated_at": now, "updated_by": _user_label(current_user)}
    _admin_track_fields = (
        "title", "description", "background", "purpose",
        "desired_due_date", "desired_deploy_date",
        "priority", "impact_scope", "is_urgent", "urgent_reason",
        "related_system", "related_menu", "related_url",
        "completion_criteria", "note",
    )
    patch_data = body.model_dump(exclude_none=True)
    patch_data.pop("submit", None)
    for field, value in patch_data.items():
        old_val = doc.get(field)
        updates[field] = value
        if field in _admin_track_fields and str(old_val) != str(value):
            await record_sr_history(sr_id, f"FIELD_CHANGE:{field}", str(old_val), str(value), _user_label(current_user))

    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one({"_id": ObjectId(sr_id)}, {"$set": updates})
    updated = await col.find_one({"_id": ObjectId(sr_id)})
    return SROut(**sr_to_out(updated))


# ── SR 검토 ───────────────────────────────────────────────────────────

@router.post("/{sr_id}/review", response_model=SROut)
async def review_sr(
    sr_id: str,
    body: SRReview,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_manager(current_user)
    doc = await get_sr_or_404(sr_id)

    if doc["status"] not in ("SUBMITTED", "REVIEWING", "PENDING_INFO"):
        raise HTTPException(status_code=400, detail="검토 가능한 상태가 아닙니다.")

    if body.result == "REJECTED" and not body.reject_reason:
        raise HTTPException(status_code=400, detail="반려 사유를 입력해야 합니다.")
    if body.result == "ON_HOLD" and not body.hold_reason:
        raise HTTPException(status_code=400, detail="보류 사유를 입력해야 합니다.")
    if body.result == "PENDING_INFO" and not body.pending_info_content:
        raise HTTPException(status_code=400, detail="추가 확인 요청 내용을 입력해야 합니다.")

    status_map = {
        "APPROVED": "APPROVED",
        "REJECTED": "REJECTED",
        "ON_HOLD": "ON_HOLD",
        "PENDING_INFO": "PENDING_INFO",
    }
    new_status = status_map[body.result]
    now = datetime.now(timezone.utc)

    updates: dict = {
        "status": new_status,
        "review_result": body.result,
        "reviewer_id": ObjectId(current_user.id),
        "reviewer_user_name": _user_label(current_user),
        "reviewed_at": now,
        "review_comment": body.comment,
        "reject_reason": body.reject_reason,
        "hold_reason": body.hold_reason,
        "pending_info_content": body.pending_info_content,
        "updated_at": now,
        "updated_by": _user_label(current_user),
    }
    if body.related_project_id:
        updates["related_project_id"] = ObjectId(body.related_project_id)
    if body.related_issue_id:
        updates["related_issue_id"] = ObjectId(body.related_issue_id)

    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one({"_id": ObjectId(sr_id)}, {"$set": updates})
    await record_status_history(sr_id, doc["status"], new_status, body.comment, _user_label(current_user))

    updated = await col.find_one({"_id": ObjectId(sr_id)})

    _STATUS_KO = {
        "APPROVED": "승인", "REJECTED": "반려", "ON_HOLD": "보류",
        "PENDING_INFO": "추가 확인 요청",
    }
    requester_id = str(doc["requester_id"])
    sender = _user_label(current_user)
    status_ko = _STATUS_KO.get(new_status, new_status)
    await create_notification(
        recipient_user_id=requester_id,
        notification_type="STATUS_CHANGED",
        title=f"SR 검토 완료: {status_ko}",
        message=f"'{doc.get('title', '')}' SR이 {status_ko} 처리되었습니다.",
        sender_user_id=str(current_user.id),
        sender_name=sender,
        target_type="SR",
        target_id=sr_id,
        target_url=f"/pm/sr/{sr_id}",
    )

    return SROut(**sr_to_out(updated))


# ── 담당자 배정 ───────────────────────────────────────────────────────

@router.post("/{sr_id}/assign", response_model=SROut)
async def assign_sr(
    sr_id: str,
    body: SRAssign,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_manager(current_user)
    doc = await get_sr_or_404(sr_id)

    if doc["status"] not in ("APPROVED", "ASSIGNED", "IN_PROGRESS"):
        raise HTTPException(status_code=400, detail="승인된 SR에만 담당자를 배정할 수 있습니다.")

    now = datetime.now(timezone.utc)
    old_assignee = doc.get("assignee_name")

    updates: dict = {
        "status": "IN_PROGRESS",
        "assignee_id": ObjectId(body.assignee_id),
        "assignee_name": body.assignee_name,
        "deployment_required": body.deployment_required,
        "security_review_required": body.security_review_required,
        "updated_at": now,
        "updated_by": _user_label(current_user),
    }
    if body.planned_start_date:
        updates["planned_start_date"] = body.planned_start_date
    if body.planned_due_date:
        updates["planned_due_date"] = body.planned_due_date
        if body.planned_due_date != doc.get("planned_due_date"):
            await record_due_date_history(
                sr_id, doc.get("planned_due_date"), body.planned_due_date, None, _user_label(current_user)
            )
    if body.estimated_effort:
        updates["estimated_effort"] = body.estimated_effort

    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one({"_id": ObjectId(sr_id)}, {"$set": updates})
    await record_sr_history(sr_id, "ASSIGNEE_CHANGE", old_assignee, body.assignee_name, _user_label(current_user))
    await record_status_history(sr_id, doc["status"], "IN_PROGRESS", None, _user_label(current_user))

    updated = await col.find_one({"_id": ObjectId(sr_id)})

    sender = _user_label(current_user)
    target_url = f"/pm/sr/{sr_id}"
    requester_id = str(doc["requester_id"])
    # 요청자에게 담당자 배정 알림
    await create_notification(
        recipient_user_id=requester_id,
        notification_type="STATUS_CHANGED",
        title="SR 담당자 배정",
        message=f"'{doc.get('title', '')}' SR이 {body.assignee_name}님에게 배정되었습니다.",
        sender_user_id=str(current_user.id),
        sender_name=sender,
        target_type="SR",
        target_id=sr_id,
        target_url=target_url,
    )
    # 담당자에게 배정 알림 (자신이 배정자인 경우 제외)
    if body.assignee_id != str(current_user.id):
        await create_notification(
            recipient_user_id=body.assignee_id,
            notification_type="ASSIGNED",
            title="SR 담당 배정",
            message=f"'{doc.get('title', '')}' SR의 담당자로 배정되었습니다.",
            sender_user_id=str(current_user.id),
            sender_name=sender,
            target_type="SR",
            target_id=sr_id,
            target_url=target_url,
        )

    # PM 이슈 자동 생성/담당자 업데이트
    from app.services.sr.sr_issue_bridge import auto_create_pm_issue, update_pm_issue_assignee
    existing_issue_id = doc.get("converted_issue_id")
    if existing_issue_id:
        await update_pm_issue_assignee(existing_issue_id, body.assignee_id)
    else:
        result = await auto_create_pm_issue(updated, body.assignee_id, current_user.id)
        if result:
            issue_id, project_id = result
            patch = {"converted_issue_id": issue_id, "converted_project_id": project_id}
            await col.update_one({"_id": ObjectId(sr_id)}, {"$set": patch})
            updated.update(patch)

    return SROut(**sr_to_out(updated))


# ── 상태 변경 ─────────────────────────────────────────────────────────

@router.post("/{sr_id}/status", response_model=SROut)
async def change_status(
    sr_id: str,
    body: SRStatusChange,
    current_user: UserPublic = Depends(get_current_user),
):
    doc = await get_sr_or_404(sr_id)
    old_status = doc["status"]
    new_status = body.status

    # 권한별 허용 상태 검증
    from app.services.sr.sr_service import (
        is_sr_admin, is_sr_manager, is_sr_operator,
        OPERATOR_ALLOWED_TARGETS, MANAGER_ALLOWED_TARGETS,
    )
    if is_sr_admin(current_user):
        pass  # 모든 상태 허용
    elif is_sr_manager(current_user):
        if new_status not in MANAGER_ALLOWED_TARGETS:
            raise HTTPException(status_code=403, detail=f"해당 상태로 변경할 권한이 없습니다: {new_status}")
    elif is_sr_operator(current_user):
        if new_status not in OPERATOR_ALLOWED_TARGETS:
            raise HTTPException(status_code=403, detail=f"해당 상태로 변경할 권한이 없습니다: {new_status}")
    else:
        raise HTTPException(status_code=403, detail="SR 처리자 이상의 권한이 필요합니다.")

    # 특수 조건 검증
    if new_status in ("REJECTED", "ON_HOLD", "CANCELLED") and not body.reason:
        raise HTTPException(status_code=400, detail="사유를 입력해야 합니다.")
    if new_status == "COMPLETED" and not body.process_result:
        raise HTTPException(status_code=400, detail="처리 결과를 입력해야 합니다.")

    now = datetime.now(timezone.utc)
    updates: dict = {"status": new_status, "updated_at": now, "updated_by": _user_label(current_user)}

    if body.process_result:
        updates["process_result"] = body.process_result
    if body.deployed is not None:
        updates["deployed"] = body.deployed
    if body.deployed_at:
        updates["deployed_at"] = body.deployed_at
    if body.actual_completed_at:
        updates["actual_completed_at"] = body.actual_completed_at
    elif new_status == "COMPLETED":
        updates["actual_completed_at"] = now
    if body.requester_confirmed is not None:
        updates["requester_confirmed"] = body.requester_confirmed

    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one({"_id": ObjectId(sr_id)}, {"$set": updates})
    await record_status_history(sr_id, old_status, new_status, body.reason, _user_label(current_user))

    updated = await col.find_one({"_id": ObjectId(sr_id)})

    _SR_STATUS_KO = {
        "SUBMITTED": "접수", "REVIEWING": "검토 중", "PENDING_INFO": "추가 확인 요청",
        "REJECTED": "반려", "APPROVED": "승인", "ASSIGNED": "담당자 배정",
        "IN_PROGRESS": "처리 중", "COMPLETED": "처리 완료",
        "CONFIRMING": "요청자 확인 중", "CLOSED": "최종 완료",
        "ON_HOLD": "보류", "CANCELLED": "취소",
    }
    new_status_ko = _SR_STATUS_KO.get(new_status, new_status)
    sender = _user_label(current_user)
    target_url = f"/pm/sr/{sr_id}"
    notify_ids = {str(doc["requester_id"])}
    if doc.get("assignee_id"):
        notify_ids.add(str(doc["assignee_id"]))
    notify_ids.discard(str(current_user.id))
    for uid in notify_ids:
        await create_notification(
            recipient_user_id=uid,
            notification_type="STATUS_CHANGED",
            title=f"SR 상태 변경: {new_status_ko}",
            message=f"'{doc.get('title', '')}' SR 상태가 {new_status_ko}(으)로 변경되었습니다.",
            sender_user_id=str(current_user.id),
            sender_name=sender,
            target_type="SR",
            target_id=sr_id,
            target_url=target_url,
        )

    return SROut(**sr_to_out(updated))


# ── 완료목표일 변경 (manager 이상) ─────────────────────────────────────

@router.post("/{sr_id}/planned-due-date", response_model=SROut)
async def change_planned_due_date(
    sr_id: str,
    body: SRDueDateChange,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_manager(current_user)
    doc = await get_sr_or_404(sr_id)

    if doc["status"] in ("DRAFT", "CLOSED", "CANCELLED", "REJECTED"):
        raise HTTPException(status_code=400, detail="완료목표일을 변경할 수 있는 상태가 아닙니다.")

    now = datetime.now(timezone.utc)
    old_due = doc.get("planned_due_date")
    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one(
        {"_id": ObjectId(sr_id)},
        {"$set": {
            "planned_due_date": body.planned_due_date,
            "updated_at": now,
            "updated_by": _user_label(current_user),
        }},
    )
    await record_due_date_history(
        sr_id, old_due, body.planned_due_date, body.change_reason, _user_label(current_user)
    )
    await record_sr_history(
        sr_id, "FIELD_CHANGE:planned_due_date",
        str(old_due)[:10] if old_due else None,
        str(body.planned_due_date)[:10], _user_label(current_user),
    )

    # 요청자에게 완료목표일 변경 알림
    await create_notification(
        recipient_user_id=str(doc["requester_id"]),
        notification_type="STATUS_CHANGED",
        title="SR 완료목표일 변경",
        message=f"'{doc.get('title', '')}' SR의 완료목표일이 {str(body.planned_due_date)[:10]}(으)로 변경되었습니다.",
        sender_user_id=str(current_user.id),
        sender_name=_user_label(current_user),
        target_type="SR",
        target_id=sr_id,
        target_url=f"/pm/sr/{sr_id}",
    )

    updated = await col.find_one({"_id": ObjectId(sr_id)})
    return SROut(**sr_to_out(updated))


# ── 내부 이슈 전환 ────────────────────────────────────────────────────

@router.post("/{sr_id}/convert-to-issue", response_model=SROut)
async def convert_to_issue(
    sr_id: str,
    issue_id: str = Query(...),
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_manager(current_user)
    doc = await get_sr_or_404(sr_id)
    now = datetime.now(timezone.utc)
    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one(
        {"_id": ObjectId(sr_id)},
        {"$set": {"converted_issue_id": issue_id, "updated_at": now, "updated_by": _user_label(current_user)}},
    )
    await record_sr_history(sr_id, "CONVERT_TO_ISSUE", None, issue_id, _user_label(current_user))
    updated = await col.find_one({"_id": ObjectId(sr_id)})
    return SROut(**sr_to_out(updated))


# ── SR 삭제 (soft delete) ─────────────────────────────────────────────

@router.delete("/{sr_id}", status_code=204)
async def delete_sr(
    sr_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_admin(current_user)
    await get_sr_or_404(sr_id)
    now = datetime.now(timezone.utc)
    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    await col.update_one(
        {"_id": ObjectId(sr_id)},
        {"$set": {"deleted_at": now, "updated_at": now, "updated_by": _user_label(current_user)}},
    )


# ── 통계 ─────────────────────────────────────────────────────────────

@router.get("/stats/summary", response_model=SRStats)
async def get_stats(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_operator(current_user)
    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]

    q: dict = {"deleted_at": None}
    if date_from or date_to:
        df: dict = {}
        if date_from:
            df["$gte"] = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        if date_to:
            df["$lte"] = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc)
        q["created_at"] = df

    docs = await col.find(q).to_list(None)

    stats = SRStats(
        total=len(docs),
        submitted=0, in_progress=0, completed=0,
        rejected=0, on_hold=0, delayed=0, cancelled=0, urgent_count=0,
    )
    processing_times: list[float] = []

    for d in docs:
        s = d.get("status", "")
        if s in ("SUBMITTED", "REVIEWING", "PENDING_INFO", "APPROVED", "ASSIGNED"):
            stats.submitted += 1
        if s in ("IN_PROGRESS", "COMPLETED", "CONFIRMING"):
            stats.in_progress += 1
        if s == "CLOSED":
            stats.completed += 1
        if s == "REJECTED":
            stats.rejected += 1
        if s == "ON_HOLD":
            stats.on_hold += 1
        if s == "CANCELLED":
            stats.cancelled += 1
        if d.get("is_urgent"):
            stats.urgent_count += 1
        if compute_is_delayed(d):
            stats.delayed += 1

        rt = d.get("request_type", "ETC")
        stats.by_type[rt] = stats.by_type.get(rt, 0) + 1

        dept = d.get("requester_department", "미지정")
        stats.by_department[dept] = stats.by_department.get(dept, 0) + 1

        sys_name = d.get("related_system") or "미지정"
        stats.by_system[sys_name] = stats.by_system.get(sys_name, 0) + 1

        assignee = d.get("assignee_name") or "미배정"
        stats.by_assignee[assignee] = stats.by_assignee.get(assignee, 0) + 1

        if d.get("actual_completed_at") and d.get("created_at"):
            delta = (d["actual_completed_at"] - d["created_at"]).total_seconds() / 86400
            processing_times.append(delta)

    if processing_times:
        stats.avg_processing_days = round(sum(processing_times) / len(processing_times), 1)

    closed_docs = [d for d in docs if d.get("status") == "CLOSED"]
    on_time = 0
    for d in closed_docs:
        completed_at = d.get("actual_completed_at")
        planned_due = d.get("planned_due_date") or d.get("desired_due_date")
        if completed_at and planned_due:
            if completed_at.tzinfo is None:
                completed_at = completed_at.replace(tzinfo=timezone.utc)
            if planned_due.tzinfo is None:
                planned_due = planned_due.replace(tzinfo=timezone.utc)
            if completed_at <= planned_due:
                on_time += 1
    if closed_docs:
        stats.on_time_rate = round(on_time / len(closed_docs) * 100, 1)

    return stats


# ── Excel 다운로드 ────────────────────────────────────────────────────

@router.get("/export", response_class=StreamingResponse)
async def export_excel(
    status: Optional[str] = Query(None),
    request_type: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    is_delayed: Optional[bool] = Query(None),
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_admin(current_user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date

    col = MongoClientManager.get_db()[MongoClientManager.SERVICE_REQUESTS]
    q: dict = {"deleted_at": None}
    if status:
        q["status"] = status
    if request_type:
        q["request_type"] = request_type
    if date_from or date_to:
        df: dict = {}
        if date_from:
            df["$gte"] = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        if date_to:
            df["$lte"] = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc)
        q["created_at"] = df

    all_docs = await col.find(q).sort("created_at", -1).to_list(None)
    if is_delayed is not None:
        docs = [d for d in all_docs if compute_is_delayed(d) == is_delayed]
    else:
        docs = all_docs

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SR목록"

    headers = [
        "SR번호", "요청제목", "요청부서", "요청자", "요청유형", "관련시스템",
        "중요도", "긴급여부", "희망완료일", "완료목표일", "실제완료일",
        "담당자", "상태", "지연여부", "접수일", "최종수정일",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    def _fmt_dt(dt):
        if not dt:
            return ""
        if hasattr(dt, "strftime"):
            return dt.strftime("%Y-%m-%d")
        return str(dt)[:10]

    for row_idx, d in enumerate(docs, 2):
        is_delayed = "Y" if compute_is_delayed(d) else "N"
        ws.append([
            d.get("sr_no", ""),
            d.get("title", ""),
            d.get("requester_department", ""),
            d.get("requester_name", ""),
            REQUEST_TYPE_LABEL.get(d.get("request_type", ""), d.get("request_type", "")),
            d.get("related_system", ""),
            SR_PRIORITY_LABEL.get(d.get("priority", ""), ""),
            "Y" if d.get("is_urgent") else "N",
            _fmt_dt(d.get("desired_due_date")),
            _fmt_dt(d.get("planned_due_date")),
            _fmt_dt(d.get("actual_completed_at")),
            d.get("assignee_name", ""),
            SR_STATUS_LABEL.get(d.get("status", ""), d.get("status", "")),
            is_delayed,
            _fmt_dt(d.get("created_at")),
            _fmt_dt(d.get("updated_at")),
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y-%m-%d")
    filename = f"SR목록_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{sr_id}/export", response_class=StreamingResponse)
async def export_sr_detail(
    sr_id: str,
    current_user: UserPublic = Depends(get_current_user),
):
    require_sr_admin(current_user)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    doc = await get_sr_or_404(sr_id)
    col_c = MongoClientManager.get_db()[MongoClientManager.SR_COMMENTS]
    col_h = MongoClientManager.get_db()[MongoClientManager.SR_STATUS_HISTORIES]
    col_fh = MongoClientManager.get_db()[MongoClientManager.SR_HISTORIES]

    comments = await col_c.find({"sr_id": sr_id, "deleted_at": None}).sort("created_at", 1).to_list(None)
    histories = await col_h.find({"sr_id": sr_id}).sort("changed_at", 1).to_list(None)
    field_histories = await col_fh.find({"sr_id": sr_id}).sort("changed_at", 1).to_list(None)

    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "요청정보"

    def _row(label, value):
        ws_info.append([label, str(value) if value is not None else ""])

    _row("SR 번호", doc.get("sr_no"))
    _row("제목", doc.get("title"))
    _row("상태", SR_STATUS_LABEL.get(doc.get("status", ""), ""))
    _row("요청자", doc.get("requester_name"))
    _row("요청부서", doc.get("requester_department"))
    _row("요청유형", REQUEST_TYPE_LABEL.get(doc.get("request_type", ""), ""))
    _row("관련시스템", doc.get("related_system"))
    _row("요청내용", doc.get("description"))
    _row("요청목적", doc.get("purpose"))
    _row("요청배경", doc.get("background"))
    _row("희망완료일", str(doc.get("desired_due_date", ""))[:10])
    _row("처리예정완료일", str(doc.get("planned_due_date", ""))[:10])
    _row("실제완료일", str(doc.get("actual_completed_at", ""))[:10])
    _row("담당자", doc.get("assignee_name"))
    _row("처리결과", doc.get("process_result"))
    _row("검토결과", doc.get("review_result"))
    _row("검토의견", doc.get("review_comment"))

    ws_c = wb.create_sheet("댓글")
    ws_c.append(["작성자", "내용", "내부메모", "작성일"])
    for c in comments:
        ws_c.append([
            c.get("writer_name", ""),
            c.get("content", ""),
            "Y" if c.get("is_internal") else "N",
            str(c.get("created_at", ""))[:19],
        ])

    ws_h = wb.create_sheet("상태이력")
    ws_h.append(["이전상태", "변경상태", "사유", "변경자", "변경일시"])
    for h in histories:
        ws_h.append([
            SR_STATUS_LABEL.get(h.get("previous_status", ""), h.get("previous_status", "")),
            SR_STATUS_LABEL.get(h.get("new_status", ""), h.get("new_status", "")),
            h.get("reason", ""),
            h.get("changed_by", ""),
            str(h.get("changed_at", ""))[:19],
        ])

    _field_label_map = {
        "title": "제목", "description": "요청내용", "background": "배경",
        "purpose": "목적", "desired_deploy_date": "희망배포일",
        "priority": "우선순위", "impact_scope": "영향범위",
        "is_urgent": "긴급여부", "urgent_reason": "긴급사유",
        "related_system": "대상시스템", "related_menu": "관련메뉴",
        "related_url": "관련URL", "completion_criteria": "완료기준", "note": "비고",
        "planned_due_date": "완료목표일",
    }
    ws_fh = wb.create_sheet("필드변경이력")
    ws_fh.append(["변경항목", "이전값", "변경값", "변경자", "변경일시"])
    for fh in field_histories:
        field_key = fh.get("action_type", "").replace("FIELD_CHANGE:", "")
        ws_fh.append([
            _field_label_map.get(field_key, field_key),
            fh.get("before_value", ""),
            fh.get("after_value", ""),
            fh.get("changed_by", ""),
            str(fh.get("changed_at", ""))[:19],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    sr_no = doc.get("sr_no", sr_id)
    filename = f"SR상세_{sr_no}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
