from __future__ import annotations

import io
import os
import zipfile
from urllib.parse import quote

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from app.db.mongo import MongoClientManager
from app.routers.isms_p.vulnerabilities import require_isms_p, UPLOAD_DIR
from app.models.user import UserPublic

router = APIRouter()


# ── 통계 ──────────────────────────────────────────────────────────────
@router.get("/stats")
async def get_stats(current_user: UserPublic = Depends(require_isms_p)):
    col = MongoClientManager.get_isms_vulnerabilities_collection()
    docs = await col.find({}, {"control_status": 1, "risk_level": 1, "asset_type": 1}).to_list(None)

    def _counts(field: str) -> dict:
        result: dict = {}
        for d in docs:
            key = d.get(field) or "미분류"
            result[key] = result.get(key, 0) + 1
        return result

    status_counts = _counts("control_status")
    risk_counts = _counts("risk_level")
    asset_counts_raw = _counts("asset_type")
    asset_counts = dict(sorted(asset_counts_raw.items(), key=lambda kv: -kv[1]))

    return {
        "total": len(docs),
        "status_counts": status_counts,
        "asset_counts": asset_counts,
        "risk_counts": risk_counts,
    }


@router.get("/assignee-stats")
async def get_assignee_stats(current_user: UserPublic = Depends(require_isms_p)):
    col = MongoClientManager.get_isms_vulnerabilities_collection()
    docs = await col.find(
        {"assignee": {"$nin": [None, ""]}, "control_status": {"$ne": "양호"}},
        {"assignee": 1, "action_status": 1, "action_difficulty": 1, "risk_level": 1},
    ).to_list(None)

    by_assignee: dict[str, dict] = {}
    for d in docs:
        a = d.get("assignee")
        if not a:
            continue
        s = by_assignee.setdefault(a, {
            "total": 0, "completed": 0,
            "hard_total": 0, "hard_completed": 0,
            "highrisk_total": 0, "highrisk_completed": 0,
        })
        s["total"] += 1
        is_done = "완료" in (d.get("action_status") or "")
        if is_done:
            s["completed"] += 1
        if d.get("action_difficulty") == "상":
            s["hard_total"] += 1
            if is_done:
                s["hard_completed"] += 1
        if d.get("risk_level") == "상":
            s["highrisk_total"] += 1
            if is_done:
                s["highrisk_completed"] += 1

    result = []
    for assignee, s in by_assignee.items():
        todo = max(0, s["total"] - s["completed"])
        completion_rate = round(s["completed"] / s["total"] * 100, 1) if s["total"] else 0.0
        hard_rate = round(s["hard_completed"] / s["hard_total"] * 100, 1) if s["hard_total"] else 0.0
        result.append({
            "assignee": assignee,
            "total": s["total"],
            "completed": s["completed"],
            "todo": todo,
            "completion_rate": completion_rate,
            "hard_total": s["hard_total"],
            "hard_completed": s["hard_completed"],
            "hard_rate": hard_rate,
            "highrisk_total": s["highrisk_total"],
            "highrisk_completed": s["highrisk_completed"],
        })
    return result


@router.get("/action-progress")
async def get_action_progress(current_user: UserPublic = Depends(require_isms_p)):
    col = MongoClientManager.get_isms_vulnerabilities_collection()
    docs = await col.find(
        {"planned_date": {"$nin": [None, ""]}, "control_status": {"$ne": "양호"}},
        {"planned_date": 1, "action_status": 1},
    ).to_list(None)

    by_month: dict[str, dict] = {}
    for d in docs:
        pd = d.get("planned_date") or ""
        month = pd[:7]
        if not month:
            continue
        m = by_month.setdefault(month, {"target": 0, "actual": 0})
        m["target"] += 1
        if d.get("action_status") == "완료":
            m["actual"] += 1

    labels = sorted(by_month.keys())
    return {
        "labels": labels,
        "target": [by_month[m]["target"] for m in labels],
        "actual": [by_month[m]["actual"] for m in labels],
    }


# ── Excel + 이미지 ZIP 내보내기 ──────────────────────────────────────
_EXPORT_COLUMNS = [
    ("id", "ID"),
    ("check_date", "점검일시"),
    ("asset_category", "자산구분"),
    ("asset_type", "자산종류"),
    ("zone", "Zone"),
    ("asset_name", "자산명"),
    ("hostname", "호스트명"),
    ("ip_address", "IP"),
    ("classification", "분류"),
    ("check_code", "점검코드"),
    ("check_item", "점검항목"),
    ("risk_level", "위험도"),
    ("check_result", "점검결과"),
    ("assignee", "담당자"),
    ("control_status", "통제여부"),
    ("action_plan", "조치계획"),
    ("planned_date", "조치예정일"),
    ("action_status", "조치여부"),
    ("action_details", "조치내용"),
    ("before_text", "수정전설명"),
    ("after_text", "수정후설명"),
    ("notes", "비고"),
]

_RISK_COLORS = {"상": "FFCCCC", "중": "FFE0B2", "하": "FFF9C4"}
_CONTROL_COLORS = {"양호": "C8E6C9", "취약": "FFCCCC", "리뷰": "FFE0B2", "해당없음": "EEEEEE"}


@router.get("/export")
async def export_vulnerabilities(current_user: UserPublic = Depends(require_isms_p)):
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    col = MongoClientManager.get_isms_vulnerabilities_collection()
    docs = await col.find({}).sort("_id", 1).to_list(None)

    by_sheet: dict[str, list[dict]] = {}
    for d in docs:
        sheet = d.get("source_sheet") or "manual"
        by_sheet.setdefault(sheet, []).append(d)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, rows in by_sheet.items():
        safe_title = sheet_name[:31] or "Sheet"
        ws = wb.create_sheet(title=safe_title)

        for ci, (_, label) in enumerate(_EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = 14 if label != "ID" else 10

        for ri, doc in enumerate(rows, start=2):
            row_height = 15
            for ci, (field, _label) in enumerate(_EXPORT_COLUMNS, start=1):
                value = str(doc["_id"]) if field == "id" else (doc.get(field) or "")
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if field == "risk_level" and value in _RISK_COLORS:
                    cell.fill = PatternFill("solid", fgColor=_RISK_COLORS[value])
                elif field == "control_status" and value in _CONTROL_COLORS:
                    cell.fill = PatternFill("solid", fgColor=_CONTROL_COLORS[value])
                elif field in ("before_text", "after_text"):
                    files_key = "before_files" if field == "before_text" else "after_files"
                    files = doc.get(files_key) or []
                    if files:
                        vuln_id = str(doc["_id"])
                        file_type = "before" if field == "before_text" else "after"
                        img_path = os.path.join(UPLOAD_DIR, vuln_id, file_type, files[0]["name"])
                        if os.path.isfile(img_path):
                            try:
                                pil_img = PILImage.open(img_path)
                                w, h = pil_img.size
                                scale = min(200 / w, 150 / h, 1.0)
                                new_w, new_h = int(w * scale), int(h * scale)
                                buf = io.BytesIO()
                                pil_img.convert("RGB").resize((new_w, new_h)).save(buf, format="PNG")
                                buf.seek(0)
                                xl_img = XLImage(buf)
                                xl_img.width, xl_img.height = new_w, new_h
                                ws.add_image(xl_img, f"{get_column_letter(ci)}{ri}")
                                row_height = max(row_height, new_h * 0.75)
                            except Exception:  # noqa: BLE001
                                pass

            ws.row_dimensions[ri].height = row_height

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_EXPORT_COLUMNS))}1"

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("취약점관리.xlsx", xlsx_buf.read())
        if os.path.isdir(UPLOAD_DIR):
            for vuln_dir in os.listdir(UPLOAD_DIR):
                vuln_path = os.path.join(UPLOAD_DIR, vuln_dir)
                if not os.path.isdir(vuln_path):
                    continue
                for file_type in ("before", "after"):
                    type_path = os.path.join(vuln_path, file_type)
                    if not os.path.isdir(type_path):
                        continue
                    for fname in os.listdir(type_path):
                        zf.write(os.path.join(type_path, fname), f"images/{vuln_dir}/{file_type}/{fname}")

    zip_buf.seek(0)
    filename = "취약점관리_export.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
