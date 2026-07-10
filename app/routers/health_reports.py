"""서버 점검(월1회) — Excel 업로드 및 보고서 조회"""
import io
import re
from datetime import datetime, timezone
from typing import Any, Optional

import openpyxl
from bson import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel

from app.db.mongo import MongoClientManager
from app.models.user import UserPublic
from app.routers.auth import get_current_user
from app.utils.mongo import oid as parse_oid

router = APIRouter()


# ── Pydantic models ──────────────────────────────────────────────────────────

class SummaryRow(BaseModel):
    no: Optional[int] = None
    host_name: str = ""
    ip: str = ""
    cpu: str = ""
    ram: str = ""
    swap: str = ""
    disk_max: str = ""
    log_errors: str = ""
    action_items: str = ""


class DiskEntry(BaseModel):
    filesystem: str = ""
    used: str = ""
    total: str = ""
    pct: str = ""


class HwCheck(BaseModel):
    item: str = ""
    ok: str = ""
    ng: str = ""
    na: str = ""


class SecurityCheck(BaseModel):
    item: str = ""
    result: str = ""


class PerfItem(BaseModel):
    before_val: str = ""
    before_pct: str = ""
    after_val: str = ""
    after_pct: str = ""


class ServerDetail(BaseModel):
    host_name: str = ""
    server_name: str = ""
    server_os: str = ""
    ip: str = ""
    inspector: str = ""
    inspection_start: str = ""
    inspection_end: str = ""
    server_shutdown: str = ""
    server_restart: str = ""
    hw_checks: list[HwCheck] = []
    cpu: PerfItem = PerfItem()
    ram: PerfItem = PerfItem()
    swap: PerfItem = PerfItem()
    network_before: str = ""
    network_after: str = ""
    disks: list[DiskEntry] = []
    security_checks: list[SecurityCheck] = []
    services: list[str] = []
    allowed_ips: str = ""
    overall_comment: str = ""


class HealthReportListItem(BaseModel):
    id: str
    report_date: str
    report_title: str
    server_count: int
    uploaded_at: Optional[str] = None
    uploaded_by: str = ""


class HealthReportOut(BaseModel):
    id: str
    report_date: str
    report_title: str
    summary: list[SummaryRow]
    servers: list[ServerDetail]
    uploaded_at: Optional[str] = None
    uploaded_by: str = ""


class HistoryPoint(BaseModel):
    report_date: str
    cpu_pct: float
    ram_pct: float
    disk_pct: float


# ── helpers ───────────────────────────────────────────────────────────────────

def _fmt_dt(dt: Any) -> Optional[str]:
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(dt)


def _s(v: Any) -> str:
    return str(v) if v is not None else ""


def _to_list_item(doc: dict) -> HealthReportListItem:
    return HealthReportListItem(
        id=str(doc["_id"]),
        report_date=doc.get("report_date", ""),
        report_title=doc.get("report_title", ""),
        server_count=len(doc.get("summary", [])),
        uploaded_at=_fmt_dt(doc.get("uploaded_at")),
        uploaded_by=doc.get("uploaded_by", ""),
    )


def _to_out(doc: dict) -> HealthReportOut:
    return HealthReportOut(
        id=str(doc["_id"]),
        report_date=doc.get("report_date", ""),
        report_title=doc.get("report_title", ""),
        summary=[SummaryRow(**s) for s in doc.get("summary", [])],
        servers=[ServerDetail(**s) for s in doc.get("servers", [])],
        uploaded_at=_fmt_dt(doc.get("uploaded_at")),
        uploaded_by=doc.get("uploaded_by", ""),
    )


# ── Excel parsers ─────────────────────────────────────────────────────────────

def _parse_summary_sheet(ws) -> tuple[str, str, list[dict]]:
    """요약 시트 → (report_date, title, summary_rows)"""
    rows = list(ws.iter_rows(values_only=True))
    title = _s(rows[0][0]) if rows else ""
    m = re.search(r"(\d{4}-\d{2}-\d{2})", title)
    report_date = m.group(1) if m else datetime.now(timezone.utc).strftime("%Y-%m-%d")

    summary = []
    for row in rows[2:]:  # skip title + header
        if not row or row[0] is None:
            continue
        try:
            no = int(row[0])
        except (TypeError, ValueError):
            continue
        log_errors = row[7]
        summary.append({
            "no": no,
            "host_name": _s(row[1]),
            "ip": _s(row[2]),
            "cpu": _s(row[3]),
            "ram": _s(row[4]),
            "swap": _s(row[5]),
            "disk_max": _s(row[6]),
            "log_errors": "-" if log_errors in (None, "-") else str(log_errors),
            "action_items": _s(row[8]),
        })
    return report_date, title, summary


def _parse_server_sheet(ws) -> dict:
    """서버 개별 시트 → ServerDetail dict"""
    rows = list(ws.iter_rows(values_only=True))

    def c(r: int, col: int) -> str:
        """1-indexed row/col → string"""
        try:
            v = rows[r - 1][col - 1]
            return str(v) if v is not None else ""
        except IndexError:
            return ""

    def t(r: int, col: int) -> str:
        """1-indexed row/col → HH:MM time string"""
        try:
            v = rows[r - 1][col - 1]
            if v is None:
                return ""
            if isinstance(v, datetime):
                return v.strftime("%H:%M")
            if hasattr(v, "hour"):  # datetime.time
                return f"{v.hour:02d}:{v.minute:02d}"
            return str(v)
        except IndexError:
            return ""

    def raw(r: int, col: int) -> Any:
        try:
            return rows[r - 1][col - 1]
        except IndexError:
            return None

    detail: dict = {}

    # ── 헤더 ──
    def t_clean(r: int, col: int) -> str:
        v = t(r, col)
        # 숫자 없이 라벨 패턴이면 빈 값으로 처리
        if v and not re.search(r"\d", v) and re.search(r"시간|종료|재기동|재가동|점검", v):
            return ""
        return v

    detail["inspector"] = c(2, 4)
    detail["inspection_start"] = t_clean(3, 2)
    detail["inspection_end"] = t_clean(4, 2)
    detail["server_shutdown"] = t_clean(3, 4)
    detail["server_restart"] = t_clean(4, 4)

    # ── 시스템 기본 정보 ──
    detail["host_name"] = c(7, 2)
    detail["server_name"] = c(8, 2)
    detail["server_os"] = c(9, 2)
    detail["ip"] = c(10, 2)

    # ── H/W 육안 점검 ──
    hw_items = ["서버 청결 상태", "파손 여부 상태", "서버 발열 상태", "케이블 정돈 상태"]
    detail["hw_checks"] = [
        {"item": item, "ok": c(7 + i, 8), "ng": c(7 + i, 9), "na": c(7 + i, 10)}
        for i, item in enumerate(hw_items)
    ]

    # ── 성능 상태 ──
    detail["cpu"] = {
        "before_val": c(14, 3), "before_pct": c(14, 5),
        "after_val": c(15, 3), "after_pct": c(15, 5),
    }
    detail["ram"] = {
        "before_val": c(16, 3), "before_pct": c(16, 5),
        "after_val": c(17, 3), "after_pct": c(17, 5),
    }
    detail["swap"] = {
        "before_val": c(18, 3), "before_pct": c(18, 5),
        "after_val": c(19, 3), "after_pct": c(19, 5),
    }
    detail["network_before"] = c(20, 3)
    detail["network_after"] = c(21, 3)

    # ── 하드웨어(Disk) — rows 14~22, cols G-J ──
    disks = []
    for ri in range(13, min(30, len(rows))):
        fs = raw(ri + 1, 7)
        if fs is None:
            continue
        fs_str = str(fs)
        used = c(ri + 1, 8)
        total = c(ri + 1, 9)
        pct = c(ri + 1, 10)
        disks.append({"filesystem": fs_str, "used": used, "total": total, "pct": pct})
        if fs_str == "Total":
            break
    detail["disks"] = disks

    # ── 시스템 보안 — rows 26-29, col G ──
    sec_items = [
        "시스템 로그, 이벤트 뷰어 확인",
        "접속로그 확인",
        "시스템 부팅 메시지 확인",
        "서버 시간(NTP) 동기화 적용 확인",
    ]
    detail["security_checks"] = [
        {"item": item, "result": c(26 + i, 7)}
        for i, item in enumerate(sec_items)
    ]

    # ── 서비스 상태 — row 33+, col B until ※ in col A ──
    services = []
    for ri in range(32, len(rows)):
        col_a = raw(ri + 1, 1)
        if col_a and str(col_a).startswith("※"):
            break
        col_b = raw(ri + 1, 2)
        if col_b:
            services.append(str(col_b))
    detail["services"] = services

    # ── ※ 접근 가능 IP / ※ 종합의견 ──
    mode = None
    allowed_lines: list[str] = []
    comment_lines: list[str] = []
    for ri in range(len(rows)):
        v = raw(ri + 1, 1)
        if v is None:
            continue
        vs = str(v)
        if vs.startswith("※ 접근 가능 IP"):
            mode = "ip"
            continue
        if vs.startswith("※ 종합의견"):
            mode = "comment"
            continue
        if mode == "ip" and vs.strip():
            allowed_lines.append(vs)
        elif mode == "comment" and vs.strip():
            comment_lines.append(vs)

    detail["allowed_ips"] = "\n".join(allowed_lines)
    detail["overall_comment"] = "\n".join(comment_lines)

    return detail


def _parse_excel(contents: bytes) -> tuple[str, str, list[dict], list[dict]]:
    """Excel 전체 파싱 → (report_date, title, summary, servers)"""
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    if "요약" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="'요약' 시트가 없습니다.")

    report_date, title, summary = _parse_summary_sheet(wb["요약"])

    # 요약 + 수동점검net_backup 제외한 나머지 시트가 서버 개별 시트
    skip = {"요약"}
    servers = []
    for sheet_name in wb.sheetnames:
        if sheet_name in skip:
            continue
        detail = _parse_server_sheet(wb[sheet_name])
        if detail.get("host_name"):
            servers.append(detail)

    return report_date, title, summary, servers


# ── routes ───────────────────────────────────────────────────────────────────

def _pct(val: Any) -> float:
    try:
        return float(str(val).replace("%", "").strip())
    except (ValueError, TypeError):
        return 0.0


@router.get("/danger")
async def get_danger_summary(current_user: UserPublic = Depends(get_current_user)):
    """최신 서버 점검 보고서에서 RAM 또는 Disk가 80% 이상인 서버 목록 반환"""
    col = MongoClientManager.get_db()[MongoClientManager.HEALTH_REPORTS]
    doc = await col.find_one({}, {"summary": 1, "report_date": 1}, sort=[("report_date", -1)])
    if not doc:
        return {"report_date": None, "servers": []}

    danger = []
    for row in (doc.get("summary") or []):
        ram_pct = _pct(row.get("ram", ""))
        disk_pct = _pct(row.get("disk_max", ""))
        if ram_pct >= 80 or disk_pct >= 80:
            danger.append({
                "host_name": row.get("host_name", ""),
                "ip": row.get("ip", ""),
                "ram": row.get("ram", ""),
                "disk_max": row.get("disk_max", ""),
                "ram_pct": ram_pct,
                "disk_pct": disk_pct,
            })

    danger.sort(key=lambda x: max(x["ram_pct"], x["disk_pct"]), reverse=True)
    return {"report_date": doc.get("report_date"), "servers": danger}


@router.get("/history/{host_name}", response_model=list[HistoryPoint])
async def get_host_history(host_name: str, current_user: UserPublic = Depends(get_current_user)):
    """특정 호스트의 월별 점검 보고서에서 RAM/Disk 사용률 추이를 반환 (오래된 순)"""
    col = MongoClientManager.get_db()[MongoClientManager.HEALTH_REPORTS]
    docs = await col.find({}, {"summary": 1, "report_date": 1}).sort("report_date", 1).to_list(None)

    target = host_name.strip().lower()
    points: list[HistoryPoint] = []
    for doc in docs:
        row = next(
            (r for r in (doc.get("summary") or []) if str(r.get("host_name", "")).strip().lower() == target),
            None,
        )
        if not row:
            continue
        points.append(HistoryPoint(
            report_date=doc.get("report_date", ""),
            cpu_pct=_pct(row.get("cpu", "")),
            ram_pct=_pct(row.get("ram", "")),
            disk_pct=_pct(row.get("disk_max", "")),
        ))
    return points


@router.get("", response_model=list[HealthReportListItem])
async def list_reports(current_user: UserPublic = Depends(get_current_user)):
    col = MongoClientManager.get_db()[MongoClientManager.HEALTH_REPORTS]
    docs = await col.find({}, {"summary": 1, "report_date": 1, "report_title": 1,
                                "uploaded_at": 1, "uploaded_by": 1}).sort("report_date", -1).to_list(None)
    return [_to_list_item(d) for d in docs]


@router.post("", response_model=HealthReportOut, status_code=status.HTTP_201_CREATED)
async def upload_report(
    file: UploadFile = File(...),
    current_user: UserPublic = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=".xlsx 파일만 업로드 가능합니다.")

    contents = await file.read()
    report_date, title, summary, servers = _parse_excel(contents)

    col = MongoClientManager.get_db()[MongoClientManager.HEALTH_REPORTS]
    now = datetime.now(timezone.utc)
    actor = current_user.full_name or current_user.email

    existing = await col.find_one({"report_date": report_date})
    if existing:
        await col.update_one(
            {"_id": existing["_id"]},
            {"$set": {"report_title": title, "summary": summary, "servers": servers,
                       "uploaded_at": now, "uploaded_by": actor}},
        )
        doc = await col.find_one({"_id": existing["_id"]})
    else:
        doc = {"report_date": report_date, "report_title": title,
               "summary": summary, "servers": servers,
               "uploaded_at": now, "uploaded_by": actor}
        result = await col.insert_one(doc)
        doc["_id"] = result.inserted_id

    return _to_out(doc)


@router.get("/{report_id}", response_model=HealthReportOut)
async def get_report(report_id: str, current_user: UserPublic = Depends(get_current_user)):
    col = MongoClientManager.get_db()[MongoClientManager.HEALTH_REPORTS]
    _id = parse_oid(report_id, "Invalid report id")
    doc = await col.find_one({"_id": _id})
    if not doc:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
    return _to_out(doc)


@router.delete("/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_report(report_id: str, current_user: UserPublic = Depends(get_current_user)):
    col = MongoClientManager.get_db()[MongoClientManager.HEALTH_REPORTS]
    _id = parse_oid(report_id, "Invalid report id")
    result = await col.delete_one({"_id": _id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다.")
