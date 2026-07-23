"""ISMS-P 취약점 Excel 임포트 서비스.

원본 isms-p 툴(FastAPI+SQLite)의 import_excel.py 를 MongoDB 기반으로 포팅.
- 기본 정보(BASE_FIELDS)는 매 임포트마다 Excel 값으로 무조건 덮어씀
- 조치 정보(ACTION_FIELDS)는 Excel 셀 값이 있을 때만 덮어씀 (비어있으면 기존 사람이 입력한 값 보존)
- 동일 (점검코드, 호스트명|자산명, 점검일시) 키의 레코드는 update, 없으면 insert
- 롤백을 위해 이번 임포트에서 변경된 레코드의 "임포트 전" 스냅샷과, 새로 삽입된 레코드의 id 목록을
  import_logs 문서에 함께 저장한다 (원본의 SQLite 파일 전체 백업 방식 대신, 배치 단위 스냅샷 방식).
"""
from __future__ import annotations

import os
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from typing import Any, Optional

import openpyxl
from bson import ObjectId
from bson.errors import InvalidId

from app.db.mongo import MongoClientManager
from app.models.isms_vulnerability import BASE_FIELDS, ACTION_FIELDS

UPLOAD_DIR = "/app/uploads/isms-p"

# ── 상태값 정규화 (원본 시트마다 표기가 제각각인 것을 통일) ─────────────────
ACTION_STATUS_MAP = {
    'O': '완료', 'o': '완료', '완료': '완료', '조치완료': '완료',
    'X': '미조치', 'x': '미조치', '미완료': '미조치', '미조치': '미조치',
    '필요없음': '필요없음',
    '': None,
}

CONTROL_STATUS_MAP = {
    'GOOD': '양호', 'GOOD ': '양호', '양호': '양호',
    'VULNERABLE': '취약', 'VULNERABLE ': '취약', '취약': '취약',
    'MANUAL': '리뷰', 'MANUAL ': '리뷰', '리뷰': '리뷰',
    'N/A': '해당없음', '#N/A': '해당없음', '해당없음': '해당없음',
    'NA': '해당없음', 'na': '해당없음', 'N/A ': '해당없음',
    '미완료': '취약', '완료': '양호',
    '인터뷰': '리뷰', '수행': '리뷰',
    '': None,
}


def normalize_action_status(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return ACTION_STATUS_MAP.get(s, s if s else None)


def normalize_control_status(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return CONTROL_STATUS_MAP.get(s, s if s else None)


def normalize_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    return s if s and s != 'None' else None


def safe_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# key: 헤더명 정규화(공백/개행 제거 + 소문자), value: (필드명, 정규화함수)
HEADER_FIELD_MAP: dict[str, Any] = {
    'id':       None,   # skip (자체 ObjectId 사용)
    '점검일시': ('check_date', normalize_date),
    '자산구분': ('asset_category', safe_str),
    '자산종류': ('asset_type', safe_str),
    'zone':     ('zone', safe_str),
    '자산명':   ('asset_name', safe_str),
    '호스트명': ('hostname', safe_str),
    'ip정보':   ('ip_address', safe_str),
    'ip':       ('ip_address', safe_str),        # WEBWAS: 'IP'
    '분류':     ('classification', safe_str),
    '항목분류': ('classification', safe_str),     # WEBWAS: '항목분류'
    '점검코드': ('check_code', safe_str),
    '점검항목': ('check_item', safe_str),
    '위험도':   ('risk_level', safe_str),
    '점검결과': ('check_result', safe_str),
    '진단내역': ('check_result', safe_str),       # WEBWAS: '진단내역'
    '담당자':   ('assignee', safe_str),
    '통제여부': ('control_status', normalize_control_status),
    '결과':     ('control_status', normalize_control_status),  # WEBWAS: '결과' = 양호/취약
    '조치계획': ('action_plan', safe_str),
    '조치예정일': ('planned_date', normalize_date),
    '조치여부': ('action_status', normalize_action_status),
    '조치내용': ('action_details', safe_str),
    '수정전설명': ('before_text', safe_str),
    '수정후설명': ('after_text', safe_str),
    '비고':     ('notes', safe_str),
}

SKIP_SHEETS = {
    '조치전_Linux', '조치전_Window', 'Sheet4', 'WA-NUMBER', 'W-Number',
    '1.표지', '2.목차', '3.개요', '4.진단기준', '5.종합결과', '5.종합결과_예전꺼', '6.진단내역',
}


def _normalize_header(v: Any) -> str:
    if v is None:
        return ''
    return str(v).strip().replace(' ', '').replace('\n', '').lower()


def _make_key(data: dict) -> tuple:
    return (
        data.get('check_code') or '',
        data.get('hostname') or data.get('asset_name') or '',
        data.get('check_date') or '',
    )


async def import_all(excel_path: str, actor_email: str | None = None) -> dict:
    """Excel(export 포맷)을 읽어 upsert. 반환: {inserted, updated, total, log_id}."""
    col = MongoClientManager.get_isms_vulnerabilities_collection()
    log_col = MongoClientManager.get_isms_import_logs_collection()

    records_before = await col.count_documents({})

    existing: dict[tuple, dict] = {}
    async for doc in col.find({}):
        key = (doc.get('check_code') or '', doc.get('hostname') or doc.get('asset_name') or '', doc.get('check_date') or '')
        if key not in existing:
            existing[key] = doc

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    inserted = 0
    updated = 0
    new_docs: list[dict] = []
    snapshots_by_id: dict[str, dict] = {}

    def upsert(data: dict, source_sheet: str) -> None:
        nonlocal inserted, updated
        check_code = data.get('check_code')
        if not check_code and not data.get('check_item'):
            return

        key = _make_key(data)

        if key in existing:
            doc = existing[key]
            if '_id' in doc:
                doc_id_str = str(doc['_id'])
                if doc_id_str not in snapshots_by_id:
                    snapshots_by_id[doc_id_str] = dict(doc)
            doc['source_sheet'] = source_sheet
            for f in BASE_FIELDS:
                if f in data:
                    doc[f] = data[f]
            for f in ACTION_FIELDS:
                if f in data and data[f] is not None:
                    doc[f] = data[f]
            doc['updated_at'] = datetime.now(timezone.utc)
            updated += 1
        else:
            new_doc = {
                'source_sheet': source_sheet,
                'before_files': [],
                'after_files': [],
                'created_at': datetime.now(timezone.utc),
                'updated_at': None,
                **data,
            }
            new_docs.append(new_doc)
            existing[key] = new_doc
            inserted += 1

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        if sheet_name == '웹취약점':
            first_row_normalized = [_normalize_header(c) for c in rows[0]]
            has_standard_header = '점검코드' in first_row_normalized or 'id' in first_row_normalized
            if not has_standard_header:
                for raw in rows[1:]:
                    if not any(raw[:8]):
                        continue
                    padded = list(raw) + [None] * 10
                    check_item = safe_str(padded[3])
                    if not check_item:
                        continue
                    url = safe_str(padded[4])
                    data = {
                        'check_code': check_item,
                        'check_item': check_item,
                        'hostname': url,
                        'ip_address': url,
                        'risk_level': safe_str(padded[2]),
                        'planned_date': normalize_date(padded[5]),
                        'assignee': safe_str(padded[6]),
                        'action_details': safe_str(padded[7]),
                        'asset_type': '웹취약점',
                        'asset_category': 'Web',
                    }
                    upsert(data, sheet_name)
                continue

        header_row_idx = None
        col_map: dict[str, int] = {}
        for i, raw in enumerate(rows):
            normalized = [_normalize_header(c) for c in raw]
            if '점검코드' in normalized or 'id' in normalized:
                header_row_idx = i
                col_map = {h: idx for idx, h in enumerate(normalized) if h}
                break

        if header_row_idx is None:
            continue

        for raw in rows[header_row_idx + 1:]:
            if not any(raw):
                continue
            padded = list(raw) + [None] * 30

            data: dict[str, Any] = {}
            for norm_header, mapping in HEADER_FIELD_MAP.items():
                if mapping is None:
                    continue
                col_idx = col_map.get(norm_header)
                if col_idx is None:
                    continue
                db_field, fn = mapping
                data[db_field] = fn(padded[col_idx])

            upsert(data, sheet_name)

    inserted_ids: list[ObjectId] = []
    if new_docs:
        result = await col.insert_many(new_docs)
        inserted_ids = list(result.inserted_ids)

    # 실제 update 반영: existing 딕셔너리에 남아있는, _id가 있던(=이미 DB에 있던) 문서들 중
    # snapshots_by_id 에 기록된(=이번 런에서 실제로 수정된) 것만 replace
    by_id_str = {str(doc['_id']): doc for doc in existing.values() if doc.get('_id') is not None}
    for doc_id_str in snapshots_by_id:
        doc = by_id_str.get(doc_id_str)
        if doc is None:
            continue
        to_set = {k: v for k, v in doc.items() if k != '_id'}
        await col.replace_one({'_id': doc['_id']}, to_set)

    await _import_images(col, excel_path)

    records_after = await col.count_documents({})
    log_doc = {
        'created_at': datetime.now(timezone.utc),
        'records_before': records_before,
        'inserted': inserted,
        'updated': updated,
        'records_after': records_after,
        'uploader_email': actor_email,
        'note': None,
        'rolled_back': False,
        'snapshots': list(snapshots_by_id.values()),
        'inserted_ids': [str(i) for i in inserted_ids],
    }
    log_result = await log_col.insert_one(log_doc)

    return {
        'inserted': inserted,
        'updated': updated,
        'total': records_after,
        'log_id': str(log_result.inserted_id),
    }


async def rollback_import(log_id: str) -> dict:
    """배치 스냅샷 기반 롤백: 이번 임포트로 새로 생긴 레코드는 삭제하고,
    수정된 레코드는 임포트 전 스냅샷으로 되돌린다."""
    log_col = MongoClientManager.get_isms_import_logs_collection()
    col = MongoClientManager.get_isms_vulnerabilities_collection()

    try:
        log_doc = await log_col.find_one({'_id': ObjectId(log_id)})
    except InvalidId:
        log_doc = None
    if not log_doc:
        raise ValueError('가져오기 이력을 찾을 수 없습니다.')
    if log_doc.get('rolled_back'):
        raise ValueError('이미 롤백되었습니다.')

    restored = 0
    for snapshot in log_doc.get('snapshots', []):
        _id = snapshot.get('_id')
        if _id is None:
            continue
        to_set = {k: v for k, v in snapshot.items() if k != '_id'}
        res = await col.replace_one({'_id': _id}, to_set)
        if res.matched_count:
            restored += 1

    deleted = 0
    for id_str in log_doc.get('inserted_ids', []):
        try:
            res = await col.delete_one({'_id': ObjectId(id_str)})
        except InvalidId:
            continue
        deleted += res.deleted_count

    await log_col.update_one({'_id': log_doc['_id']}, {'$set': {'rolled_back': True}})

    return {'restored': restored, 'deleted': deleted}


# ── 임포트 시 Excel 셀에 삽입된 수정전/후 이미지 자동 추출 ──────────────────

async def _import_images(col, xlsx_path: str) -> int:
    R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    S = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    A = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    saved = 0
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = set(zf.namelist())

            wb_xml = ET.fromstring(zf.read('xl/workbook.xml'))
            wb_rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            rId_to_target = {rel.get('Id'): rel.get('Target') for rel in wb_rels}

            sheets = []
            for s in wb_xml.iter(f'{{{S}}}sheet'):
                rId = s.get(f'{{{R}}}id')
                if rId in rId_to_target:
                    t = rId_to_target[rId].lstrip('/')
                    path = t if t.startswith('xl/') else f'xl/{t}'
                    sheets.append((s.get('name'), path))

            for sheet_name, sheet_path in sheets:
                if sheet_path not in names:
                    continue
                sheet_dir, sheet_file = sheet_path.rsplit('/', 1)
                rels_path = f'{sheet_dir}/_rels/{sheet_file}.rels'
                if rels_path not in names:
                    continue

                sheet_rels = ET.fromstring(zf.read(rels_path))
                drawing_target = None
                for rel in sheet_rels:
                    if 'drawing' in rel.get('Type', '').lower():
                        drawing_target = rel.get('Target').lstrip('/')
                        break
                if not drawing_target:
                    continue

                drawing_path = os.path.normpath(f'{sheet_dir}/{drawing_target}').replace('\\', '/')
                if drawing_path not in names:
                    drawing_path = drawing_target
                if drawing_path not in names:
                    continue

                drawing_dir, drawing_file = drawing_path.rsplit('/', 1)
                drw_rels_path = f'{drawing_dir}/_rels/{drawing_file}.rels'
                if drw_rels_path not in names:
                    continue

                drw_rels = ET.fromstring(zf.read(drw_rels_path))
                rId_to_img: dict[str, tuple[str, str]] = {}
                for rel in drw_rels:
                    if 'image' in rel.get('Type', '').lower():
                        t = rel.get('Target').lstrip('/')
                        img_path = os.path.normpath(f'{drawing_dir}/{t}').replace('\\', '/')
                        if img_path not in names:
                            img_path = t
                        if img_path in names:
                            ext = img_path.rsplit('.', 1)[-1].lower() if '.' in img_path else 'png'
                            rId_to_img[rel.get('Id')] = (img_path, ext)

                if not rId_to_img:
                    continue

                cell_to_img: dict[tuple[int, int], tuple[bytes, str]] = {}
                drawing_xml = ET.fromstring(zf.read(drawing_path))
                for anchor in drawing_xml:
                    tag = anchor.tag.split('}')[-1]
                    if tag not in ('oneCellAnchor', 'twoCellAnchor'):
                        continue
                    from_elem = anchor.find(f'{{{XDR}}}from')
                    if from_elem is None:
                        continue
                    col_e = from_elem.find(f'{{{XDR}}}col')
                    row_e = from_elem.find(f'{{{XDR}}}row')
                    if col_e is None or row_e is None:
                        continue
                    c = int(col_e.text) + 1
                    r = int(row_e.text) + 1
                    blip = anchor.find(f'.//{{{A}}}blip')
                    if blip is None:
                        continue
                    rId = blip.get(f'{{{R}}}embed')
                    if rId in rId_to_img:
                        img_file, ext = rId_to_img[rId]
                        cell_to_img[(c, r)] = (zf.read(img_file), ext)

                if not cell_to_img:
                    continue

                wb2 = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
                ws2 = wb2[sheet_name]
                rows = list(ws2.iter_rows(values_only=True))
                wb2.close()

                id_col = before_col = after_col = check_code_col = hostname_col = check_date_col = header_row_idx = None
                for i, raw in enumerate(rows):
                    norm = [_normalize_header(c) for c in raw]
                    if '점검코드' in norm:
                        header_row_idx = i
                        for j, h in enumerate(norm):
                            if h == 'id':
                                id_col = j
                            elif h == '점검코드':
                                check_code_col = j
                            elif h == '호스트명':
                                hostname_col = j
                            elif h == '점검일시':
                                check_date_col = j
                            elif h == '수정전설명':
                                before_col = j
                            elif h == '수정후설명':
                                after_col = j
                        break

                if header_row_idx is None or check_code_col is None:
                    continue

                ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                for di, raw in enumerate(rows[header_row_idx + 1:]):
                    if not any(raw):
                        continue
                    padded = list(raw) + [None] * 5
                    ws_row = header_row_idx + 2 + di

                    vuln = None
                    if id_col is not None:
                        raw_id = safe_str(padded[id_col])
                        if raw_id:
                            try:
                                vuln = await col.find_one({'_id': ObjectId(raw_id)})
                            except InvalidId:
                                vuln = None
                    if vuln is None:
                        check_code = safe_str(padded[check_code_col]) if check_code_col is not None else None
                        hostname = safe_str(padded[hostname_col]) if hostname_col is not None else None
                        check_date = safe_str(padded[check_date_col]) if check_date_col is not None else None
                        if check_code:
                            q: dict[str, Any] = {'check_code': check_code}
                            if hostname:
                                q['hostname'] = hostname
                            if check_date:
                                q2 = dict(q, check_date=check_date)
                                vuln = await col.find_one(q2)
                            if vuln is None:
                                vuln = await col.find_one(q)
                    if not vuln:
                        continue

                    vuln_id_str = str(vuln['_id'])
                    for file_type, col_idx, files_attr in [
                        ('before', before_col, 'before_files'),
                        ('after', after_col, 'after_files'),
                    ]:
                        if col_idx is None:
                            continue
                        ws_col = col_idx + 1
                        img_data = cell_to_img.get((ws_col, ws_row))
                        if not img_data:
                            continue
                        img_bytes, ext = img_data
                        fname = f'imported_{ts}_{di}.{ext}'
                        dest_dir = os.path.join(UPLOAD_DIR, vuln_id_str, file_type)
                        os.makedirs(dest_dir, exist_ok=True)
                        with open(os.path.join(dest_dir, fname), 'wb') as f:
                            f.write(img_bytes)
                        await col.update_one(
                            {'_id': vuln['_id']},
                            {'$set': {files_attr: [{'name': fname, 'original': fname}]}},
                        )
                        saved += 1
    except Exception as e:  # noqa: BLE001 — 이미지 추출 실패는 임포트 전체를 막지 않음
        import logging
        logging.getLogger(__name__).warning('isms-p image extraction failed: %s', e)

    return saved
